# -*- coding: utf-8 -*-
"""财务核对核心业务模块。

本文件负责 SQLite 数据库、Excel/CSV 导入、字段映射、金额计算、
汇总查询、明细分页、差异分组和备份还原。Qt 主界面只调用 Repository，
不直接操作数据库，从而让业务逻辑可以独立测试和复用。
"""
import csv
import hashlib
import json
import os
import re
import sqlite3
import sys
import traceback
import zipfile
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


APP_TITLE = "财务第三方支付核对"
if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).resolve().parent
else:
    BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DB_PATH = DATA_DIR / "reconciliation.sqlite3"

REPORT_TYPES = ["已结算", "未结算"]

# 默认汇总字段。每个店铺后续可在参数配置中单独调整。
SUMMARY_COLUMNS = [
    "店铺", "报表类型", "月份", "订单实付应结", "平台补贴", "商家补贴", "结算运费", "订单退款",
    "收入净额合计", "已结算佣金", "技术服务费", "支出金额", "结算金额", "提现金额", "期初金额", "结算期末余额",
]

# 默认原始流水字段。导入时会根据实际表头补充每个店铺自己的字段配置。
RAW_COLUMNS = [
    "店铺", "月份", "动账时间", "动账流水号", "动账方向", "动账账户", "动账金额",
    "动账摘要", "业务类型", "主订单编号", "子订单编号", "售后单号", "下单时间",
    "商品信息", "商品编码", "售卖类型", "订单实付应结", "平台补贴", "商家补贴",
    "结算运费", "订单退款", "佣金", "技术服务费",
]

# 手工调整可以影响的汇总字段；“备查”只记录说明，不改变汇总金额。
ADJUSTABLE_COLUMNS = [
    "备查（不影响汇总）", "订单实付应结", "平台补贴", "商家补贴", "结算运费", "订单退款",
    "已结算佣金", "技术服务费", "提现金额", "期初金额",
]

SUMMARY_EXTRA_COLUMNS = ["店铺期末余额", "差异", "差异原因", "调整说明"]

DEFAULT_FORMULA_NOTE = (
    "收入净额合计=订单实付应结+平台补贴+商家补贴+结算运费+订单退款\n"
    "支出金额=已结算佣金+技术服务费\n"
    "结算金额=收入净额合计+支出金额\n"
    "结算期末余额=期初金额+收入净额合计+支出金额+提现金额"
)

FIELD_KEYS = {
    "订单实付应结": "paid_settlement",
    "平台补贴": "platform_subsidy",
    "商家补贴": "merchant_subsidy",
    "结算运费": "freight",
    "订单退款": "refund",
    "收入净额合计": "income_total",
    "已结算佣金": "commission",
    "技术服务费": "tech_fee",
    "支出金额": "expense_amount",
    "结算金额": "settlement_amount",
    "提现金额": "withdraw_amount",
    "期初金额": "opening_balance",
    "结算期末余额": "ending_balance",
}


LEADING_TEXT_MARKS = "'’＇`"


def clean_cell_text(value):
    """清理 Excel/CSV 单元格文本，去掉订单号前常见的文本标记。"""

    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    text = str(value).strip()
    return text.lstrip(LEADING_TEXT_MARKS)


def sql_literal(value):
    """把文本安全转换成 SQL 字面量，主要用于生成清洗表达式。"""

    return "'" + str(value).replace("'", "''") + "'"


def clean_sql_expr(expr):
    """生成 SQLite 表达式，查询时清理前导文本标记。"""

    return f"LTRIM(CAST({expr} AS TEXT), {sql_literal(LEADING_TEXT_MARKS)})"


def money(value) -> Decimal:
    """把输入值转换为 Decimal 金额，并统一保留两位小数。"""

    if value is None or value == "":
        return Decimal("0")
    try:
        return Decimal(clean_cell_text(value).replace(",", "").strip()).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def money_text(value) -> str:
    """把金额格式化为带千分位的显示文本。"""

    return f"{money(value):,.2f}"


def db_float(value) -> float:
    """把金额转换为 SQLite 可存储的 float。"""

    return float(money(value))


def normalize_header(value) -> str:
    """标准化表头，便于匹配不同平台的字段别名。"""

    return clean_cell_text(value).replace("\n", "").replace("\r", "")


HEADER_ALIASES = {
    "店铺": ["店铺"],
    "月份": ["动账时间1", "动帐时间1", "月份", "账期", "帐期"],
    "动账时间": ["动账时间", "动帐时间", "交易时间", "发生时间"],
    "动账流水号": ["动账流水号", "动帐流水号", "流水号", "资金流水号"],
    "动账方向": ["动账方向", "动帐方向", "收支方向"],
    "动账账户": ["动账账户", "动帐账户", "账户"],
    "动账金额": ["动账金额", "动帐金额", "金额"],
    "动账摘要": ["动账摘要", "动帐摘要", "动账场景", "动帐场景", "计费类型", "摘要"],
    "业务类型": ["业务类型", "订单类型", "计费类型"],
    "主订单编号": ["主订单编号", "订单号"],
    "子订单编号": ["子订单编号", "子订单号"],
    "售后单号": ["售后单号", "售后编号"],
    "下单时间": ["下单时间"],
    "商品信息": ["商品信息", "商品名称"],
    "商品编码": ["商品编码", "商品ID"],
    "售卖类型": ["售卖类型", "订单类型"],
}

AMOUNT_ALIASES = {
    "订单实付应结": ["订单实付应结"],
    "平台补贴": ["平台补贴", "实际平台补贴_运费", "实际平台补贴", "其他平台补贴", "政府补贴平台垫资"],
    "商家补贴": ["商家补贴", "实际达人补贴", "实际抖音支付补贴", "实际抖音月付营销补贴", "银行补贴", "以旧换新抵扣"],
    "结算运费": ["结算运费", "运费实付"],
    "订单退款": ["订单退款"],
    "佣金": ["佣金", "服务商佣金", "渠道分成", "招商服务费", "站外推广费", "其他分成"],
    "技术服务费": ["技术服务费", "平台服务费"],
}


def difference_reason(row) -> str:
    """根据余额核对结果生成一段可读的差异说明。"""

    if row["account_ending"] is None:
        return "未录入店铺期末余额"
    if row["difference"] == Decimal("0.00"):
        return "一致"
    direction = "结算期末余额大于店铺期末余额" if row["difference"] > 0 else "结算期末余额小于店铺期末余额"
    return f"{direction}；请核对收入净额合计、支出金额、提现金额、期初金额或店铺期末余额录入"


def parse_month_sort(label, transaction_time=None) -> str:
    """把各种年月显示格式统一成 YYYY-MM 排序值。"""

    label = str(label or "").strip()
    digits = "".join(ch for ch in label if ch.isdigit())
    if "月" in label and digits:
        year = "0000"
        if transaction_time:
            text = str(transaction_time)
            if len(text) >= 4 and text[:4].isdigit():
                year = text[:4]
        return f"{year}-{int(digits):02d}"
    if len(digits) == 6:
        return f"{digits[:4]}-{digits[4:]}"
    if transaction_time:
        text = str(transaction_time)
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(text[:19], fmt).strftime("%Y-%m")
            except ValueError:
                pass
        if len(text) >= 7 and text[4] in "-/":
            return text[:7].replace("/", "-")
    return label


def normalize_month_token(value) -> str:
    """把用户输入的年月片段统一成 YYYY-MM。"""

    text = str(value or "").strip()
    if not text:
        return ""
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) >= 6:
        return f"{digits[:4]}-{digits[4:6]}"
    if len(text) >= 7 and text[4] in "-/":
        return text[:7].replace("/", "-")
    return text


def parse_month_filter(value):
    """解析首页年月筛选，支持逗号分隔和连续年月区间。"""

    text = str(value or "").strip()
    if not text:
        return []
    text = text.replace("，", ",").replace("、", ",").replace("至", "-").replace("到", "-")
    if "," in text:
        return [m for m in (normalize_month_token(part) for part in text.split(",")) if m]
    if "-" in text:
        compact = text.replace(" ", "")
        digits = "".join(ch for ch in compact if ch.isdigit())
        if len(digits) == 12:
            start = normalize_month_token(digits[:6])
            end = normalize_month_token(digits[6:])
            return month_range(start, end)
    month = normalize_month_token(text)
    return [month] if month else []


def month_range(start, end):
    """生成两个年月之间的连续月份列表。"""

    start = normalize_month_token(start)
    end = normalize_month_token(end)
    if not start or not end:
        return []
    sy, sm = [int(x) for x in start.split("-")]
    ey, em = [int(x) for x in end.split("-")]
    months = []
    y, m = sy, sm
    while (y, m) <= (ey, em):
        months.append(f"{y:04d}-{m:02d}")
        m += 1
        if m == 13:
            y += 1
            m = 1
    return months


def json_default(value):
    """把日期和 Decimal 转成 JSON 可序列化的值。"""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, Decimal):
        return str(value)
    return clean_cell_text(value)


def parse_formula_lines(text):
    """解析店铺自定义计算口径，把“字段=表达式”拆成可执行规则。"""

    formulas = []
    normalized = str(text or "").replace("；", "\n").replace(";", "\n")
    for line in normalized.splitlines():
        line = line.strip().rstrip("。")
        if not line or "=" not in line:
            continue
        left, right = line.split("=", 1)
        left = left.strip()
        right = right.strip()
        if left in FIELD_KEYS and right:
            formulas.append((left, right))
    return formulas


def eval_money_formula(expr, values):
    """在受控变量范围内计算金额公式。"""

    names = sorted(FIELD_KEYS.keys(), key=len, reverse=True)
    local_vars = {}
    safe_expr = str(expr)
    for idx, name in enumerate(names):
        var = f"v{idx}"
        local_vars[var] = money(values.get(FIELD_KEYS[name], 0))
        safe_expr = safe_expr.replace(name, var)
    allowed = set("0123456789.+-*/() _v")
    if any(ch not in allowed for ch in safe_expr):
        raise ValueError(f"公式包含不支持的字符：{expr}")
    return money(eval(safe_expr, {"__builtins__": {}}, local_vars))


@dataclass
class ImportResult:
    """导入结果统计，用于展示新增行数、跳过行数和导入文件名。"""

    imported: int
    skipped: int
    file_name: str


class Repository:
    """应用的数据访问层。

    主库保存店铺清单和全局设置；每个店铺拥有独立 SQLite 文件保存流水、
    余额、调整和字段配置。所有界面查询、导入、导出和备份还原都通过
    本类完成。
    """

    def __init__(self, path: Path):
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        self.path = Path(path)
        self.store_dir = DATA_DIR / "stores"
        self.store_dir.mkdir(parents=True, exist_ok=True)
        self.master_conn = sqlite3.connect(self.path, timeout=30)
        self.master_conn.row_factory = sqlite3.Row
        self.master_conn.execute("PRAGMA busy_timeout=30000")
        self.conn = self.master_conn
        self.conn.row_factory = sqlite3.Row
        self.current_store = ""
        self.store_conns = {}
        self.initialized_store_dbs = set()
        self.init_db()
        self._ensure_master_columns()

    def init_db(self):
        """初始化主库、店铺目录和默认字段配置表。"""

        cur = self.conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS stores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                note TEXT,
                active INTEGER DEFAULT 1,
                created_at TEXT NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS app_settings (
                key TEXT PRIMARY KEY,
                value TEXT,
                updated_at TEXT NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_file TEXT,
                imported_at TEXT,
                store TEXT NOT NULL,
                report_type TEXT DEFAULT '已结算',
                month_label TEXT NOT NULL,
                month_sort TEXT NOT NULL,
                transaction_time TEXT,
                flow_id TEXT,
                direction TEXT,
                account TEXT,
                amount REAL DEFAULT 0,
                summary TEXT,
                biz_type TEXT,
                main_order TEXT,
                sub_order TEXT,
                after_sale TEXT,
                order_time TEXT,
                product_info TEXT,
                product_code TEXT,
                sale_type TEXT,
                paid_settlement REAL DEFAULT 0,
                platform_subsidy REAL DEFAULT 0,
                merchant_subsidy REAL DEFAULT 0,
                freight REAL DEFAULT 0,
                refund REAL DEFAULT 0,
                commission REAL DEFAULT 0,
                tech_fee REAL DEFAULT 0,
                raw_payload TEXT,
                UNIQUE(store, report_type, month_sort, flow_id, sub_order, summary, amount)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS store_configs (
                store TEXT PRIMARY KEY,
                raw_columns TEXT,
                summary_columns TEXT,
                frozen_columns TEXT,
                page_size INTEGER DEFAULT 1000,
                formula_note TEXT,
                updated_at TEXT NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS balances (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                store TEXT NOT NULL,
                report_type TEXT DEFAULT '已结算',
                month_label TEXT NOT NULL,
                month_sort TEXT NOT NULL,
                opening_balance REAL DEFAULT 0,
                account_ending_balance REAL,
                note TEXT,
                updated_at TEXT NOT NULL,
                UNIQUE(store, report_type, month_sort)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS adjustments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                store TEXT NOT NULL,
                report_type TEXT DEFAULT '已结算',
                month_label TEXT NOT NULL,
                month_sort TEXT NOT NULL,
                target_column TEXT DEFAULT '备查（不影响汇总）',
                item TEXT NOT NULL,
                amount REAL DEFAULT 0,
                note TEXT,
                updated_at TEXT NOT NULL
            )
        """)
        adj_cols = {row["name"] for row in cur.execute("PRAGMA table_info(adjustments)").fetchall()}
        if "target_column" not in adj_cols:
            cur.execute("ALTER TABLE adjustments ADD COLUMN target_column TEXT DEFAULT '备查（不影响汇总）'")
        tx_cols = {row["name"] for row in cur.execute("PRAGMA table_info(transactions)").fetchall()}
        if "raw_payload" not in tx_cols:
            cur.execute("ALTER TABLE transactions ADD COLUMN raw_payload TEXT")
        if "report_type" not in tx_cols:
            cur.execute("ALTER TABLE transactions ADD COLUMN report_type TEXT DEFAULT '已结算'")
            cur.execute("UPDATE transactions SET report_type='已结算' WHERE report_type IS NULL OR TRIM(report_type)=''")
        balance_cols = {row["name"] for row in cur.execute("PRAGMA table_info(balances)").fetchall()}
        if "report_type" not in balance_cols:
            cur.execute("ALTER TABLE balances ADD COLUMN report_type TEXT DEFAULT '已结算'")
            cur.execute("UPDATE balances SET report_type='已结算' WHERE report_type IS NULL OR TRIM(report_type)=''")
        adjustment_cols = {row["name"] for row in cur.execute("PRAGMA table_info(adjustments)").fetchall()}
        if "report_type" not in adjustment_cols:
            cur.execute("ALTER TABLE adjustments ADD COLUMN report_type TEXT DEFAULT '已结算'")
            cur.execute("UPDATE adjustments SET report_type='已结算' WHERE report_type IS NULL OR TRIM(report_type)=''")
        self._migrate_report_type_unique_keys(cur)
        config_cols = {row["name"] for row in cur.execute("PRAGMA table_info(store_configs)").fetchall()}
        if "frozen_columns" not in config_cols:
            cur.execute("ALTER TABLE store_configs ADD COLUMN frozen_columns TEXT")
        if "page_size" not in config_cols:
            cur.execute("ALTER TABLE store_configs ADD COLUMN page_size INTEGER DEFAULT 1000")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_transactions_store_type_month_time ON transactions(store, report_type, month_sort, transaction_time, id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_transactions_month ON transactions(month_sort)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_balances_store_type_month ON balances(store, report_type, month_sort)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_adjustments_store_type_month ON adjustments(store, report_type, month_sort)")
        cur.execute("""
            INSERT OR IGNORE INTO stores (name, note, active, created_at)
            SELECT DISTINCT store, '', 1, ?
            FROM transactions
            WHERE store IS NOT NULL AND TRIM(store) <> ''
        """, (datetime.now().strftime("%Y-%m-%d %H:%M:%S"),))
        cur.execute("""
            INSERT OR IGNORE INTO store_configs (store, raw_columns, summary_columns, frozen_columns, page_size, formula_note, updated_at)
            SELECT name, ?, ?, ?, ?, ?, ? FROM stores
        """, (
            json.dumps(RAW_COLUMNS, ensure_ascii=False),
            json.dumps(SUMMARY_COLUMNS + SUMMARY_EXTRA_COLUMNS, ensure_ascii=False),
            json.dumps(SUMMARY_COLUMNS[:3], ensure_ascii=False),
            1000,
            DEFAULT_FORMULA_NOTE,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ))
        self.conn.commit()

    def get_app_setting(self, key, default=""):
        row = self.master_conn.execute("SELECT value FROM app_settings WHERE key=?", (key,)).fetchone()
        if not row:
            return default
        return row["value"] if row["value"] is not None else default

    def set_app_setting(self, key, value):
        self.master_conn.execute("""
            INSERT INTO app_settings (key, value, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
        """, (key, str(value), datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        self.master_conn.commit()

    def app_settings(self):
        return [dict(row) for row in self.master_conn.execute("SELECT key, value, updated_at FROM app_settings ORDER BY key").fetchall()]

    def backup_data(self, backup_path, store_names=None, include_settings=True):
        """把主配置、全局设置和选中店铺数据库打包为 ZIP。"""

        store_names = [str(name).strip() for name in (store_names or []) if str(name).strip()]
        if store_names:
            placeholders = ",".join("?" for _ in store_names)
            rows = self.master_conn.execute(f"SELECT * FROM stores WHERE name IN ({placeholders}) ORDER BY name", store_names).fetchall()
        else:
            rows = self.master_conn.execute("SELECT * FROM stores ORDER BY name").fetchall()
        stores = [dict(row) for row in rows]
        manifest = {
            "app": "payment_reconciliation_qt",
            "version": "1.0.3",
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "include_settings": bool(include_settings),
            "stores": stores,
            "settings": self.app_settings() if include_settings else [],
        }
        backup_path = Path(backup_path)
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(backup_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
            for store in stores:
                db_file = store.get("db_file") or self._default_store_db_file(store.get("name", ""))
                db_path = self.store_dir / db_file
                if db_path.exists():
                    zf.write(db_path, f"stores/{db_file}")
        return {"stores": len(stores), "path": str(backup_path)}

    def inspect_backup(self, backup_path):
        with zipfile.ZipFile(backup_path, "r") as zf:
            manifest = json.loads(zf.read("manifest.json").decode("utf-8"))
        if manifest.get("app") != "payment_reconciliation_qt":
            raise ValueError("不是有效的财务核对备份文件")
        return manifest

    def _close_store_connection_for_file(self, db_file):
        db_path = self.store_dir / db_file
        key = str(db_path)
        conn = self.store_conns.pop(key, None)
        if conn:
            conn.close()
        if self.conn is conn:
            self.conn = self.master_conn
            self.current_store = ""
        self.initialized_store_dbs.discard(key)

    def restore_data(self, backup_path, store_names=None, restore_settings=True):
        """从 ZIP 还原店铺数据库和配置，还原前会关闭相关店铺连接。"""

        manifest = self.inspect_backup(backup_path)
        wanted = {str(name).strip() for name in (store_names or []) if str(name).strip()}
        stores = [store for store in manifest.get("stores", []) if not wanted or store.get("name") in wanted]
        with zipfile.ZipFile(backup_path, "r") as zf:
            names = set(zf.namelist())
            for store in stores:
                name = store.get("name", "").strip()
                if not name:
                    continue
                db_file = store.get("db_file") or self._default_store_db_file(name)
                member = f"stores/{db_file}"
                self._close_store_connection_for_file(db_file)
                if member in names:
                    target = self.store_dir / db_file
                    target.parent.mkdir(parents=True, exist_ok=True)
                    with zf.open(member, "r") as src, open(target, "wb") as dst:
                        dst.write(src.read())
                self.master_conn.execute("""
                    INSERT INTO stores (name, note, active, created_at, db_file)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(name) DO UPDATE SET
                        note=excluded.note,
                        active=excluded.active,
                        db_file=excluded.db_file
                """, (
                    name,
                    store.get("note", ""),
                    int(store.get("active", 1) or 0),
                    store.get("created_at") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    db_file,
                ))
        if restore_settings:
            for setting in manifest.get("settings", []):
                self.master_conn.execute("""
                    INSERT INTO app_settings (key, value, updated_at)
                    VALUES (?, ?, ?)
                    ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """, (
                    setting.get("key"),
                    setting.get("value"),
                    setting.get("updated_at") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                ))
        self.master_conn.commit()
        return {"stores": len(stores), "settings": len(manifest.get("settings", [])) if restore_settings else 0}

    def _unique_index_columns(self, cur, table):
        result = []
        for index in cur.execute(f"PRAGMA index_list({table})").fetchall():
            if not index["unique"]:
                continue
            cols = [row["name"] for row in cur.execute(f"PRAGMA index_info({index['name']})").fetchall()]
            result.append(cols)
        return result

    def _has_unique_columns(self, cur, table, columns):
        expected = list(columns)
        return any(cols == expected for cols in self._unique_index_columns(cur, table))

    def _migrate_report_type_unique_keys(self, cur):
        if not self._has_unique_columns(cur, "balances", ["store", "report_type", "month_sort"]):
            cur.execute("ALTER TABLE balances RENAME TO balances_old")
            cur.execute("""
                CREATE TABLE balances (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    store TEXT NOT NULL,
                    report_type TEXT DEFAULT '已结算',
                    month_label TEXT NOT NULL,
                    month_sort TEXT NOT NULL,
                    opening_balance REAL DEFAULT 0,
                    account_ending_balance REAL,
                    note TEXT,
                    updated_at TEXT NOT NULL,
                    UNIQUE(store, report_type, month_sort)
                )
            """)
            cur.execute("""
                INSERT OR IGNORE INTO balances
                (id, store, report_type, month_label, month_sort, opening_balance, account_ending_balance, note, updated_at)
                SELECT id, store, COALESCE(NULLIF(TRIM(report_type), ''), '已结算'), month_label, month_sort,
                       opening_balance, account_ending_balance, note, updated_at
                FROM balances_old
            """)
            cur.execute("DROP TABLE balances_old")
        if not self._has_unique_columns(cur, "transactions", ["store", "report_type", "month_sort", "flow_id", "sub_order", "summary", "amount"]):
            cur.execute("ALTER TABLE transactions RENAME TO transactions_old")
            cur.execute("""
                CREATE TABLE transactions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source_file TEXT,
                    imported_at TEXT,
                    store TEXT NOT NULL,
                    report_type TEXT DEFAULT '已结算',
                    month_label TEXT NOT NULL,
                    month_sort TEXT NOT NULL,
                    transaction_time TEXT,
                    flow_id TEXT,
                    direction TEXT,
                    account TEXT,
                    amount REAL DEFAULT 0,
                    summary TEXT,
                    biz_type TEXT,
                    main_order TEXT,
                    sub_order TEXT,
                    after_sale TEXT,
                    order_time TEXT,
                    product_info TEXT,
                    product_code TEXT,
                    sale_type TEXT,
                    paid_settlement REAL DEFAULT 0,
                    platform_subsidy REAL DEFAULT 0,
                    merchant_subsidy REAL DEFAULT 0,
                    freight REAL DEFAULT 0,
                    refund REAL DEFAULT 0,
                    commission REAL DEFAULT 0,
                    tech_fee REAL DEFAULT 0,
                    raw_payload TEXT,
                    UNIQUE(store, report_type, month_sort, flow_id, sub_order, summary, amount)
                )
            """)
            cur.execute("""
                INSERT OR IGNORE INTO transactions
                (id, source_file, imported_at, store, report_type, month_label, month_sort, transaction_time, flow_id,
                 direction, account, amount, summary, biz_type, main_order, sub_order, after_sale, order_time,
                 product_info, product_code, sale_type, paid_settlement, platform_subsidy, merchant_subsidy,
                 freight, refund, commission, tech_fee, raw_payload)
                SELECT id, source_file, imported_at, store, COALESCE(NULLIF(TRIM(report_type), ''), '已结算'),
                       month_label, month_sort, transaction_time, flow_id, direction, account, amount, summary,
                       biz_type, main_order, sub_order, after_sale, order_time, product_info, product_code, sale_type,
                       paid_settlement, platform_subsidy, merchant_subsidy, freight, refund, commission, tech_fee, raw_payload
                FROM transactions_old
            """)
            cur.execute("DROP TABLE transactions_old")

    def _ensure_master_columns(self):
        cur = self.master_conn.cursor()
        cols = {row["name"] for row in cur.execute("PRAGMA table_info(stores)").fetchall()}
        if "db_file" not in cols:
            cur.execute("ALTER TABLE stores ADD COLUMN db_file TEXT")
        rows = cur.execute("SELECT id, name, db_file FROM stores").fetchall()
        for row in rows:
            if not row["db_file"]:
                cur.execute("UPDATE stores SET db_file=? WHERE id=?", (self._default_store_db_file(row["name"]), row["id"]))
        self.master_conn.commit()

    def _default_store_db_file(self, store):
        safe = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff_-]+", "_", str(store or "").strip()).strip("_")
        safe = safe[:40] or "store"
        digest = hashlib.sha1(str(store or "").encode("utf-8")).hexdigest()[:8]
        return f"{safe}_{digest}.sqlite3"

    def _master_store_row(self, store):
        return self.master_conn.execute("SELECT * FROM stores WHERE name=?", (store,)).fetchone()

    def store_db_path(self, store):
        store = str(store or "").strip()
        if not store:
            return self.path
        row = self._master_store_row(store)
        if not row:
            self.master_conn.execute("""
                INSERT INTO stores (name, note, active, created_at, db_file)
                VALUES (?, '', 1, ?, ?)
            """, (store, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), self._default_store_db_file(store)))
            self.master_conn.commit()
            row = self._master_store_row(store)
        db_file = row["db_file"] or self._default_store_db_file(store)
        return self.store_dir / db_file

    def _store_db_ready(self, conn):
        try:
            rows = conn.execute("""
                SELECT name FROM sqlite_master
                WHERE type='table' AND name IN ('transactions', 'balances', 'adjustments', 'store_configs')
            """).fetchall()
            return len(rows) == 4
        except sqlite3.OperationalError:
            return False

    def use_store(self, store):
        """切换当前店铺，并确保该店铺独立库已初始化。"""

        store = str(store or "").strip()
        if not store:
            self.conn = self.master_conn
            self.current_store = ""
            return
        if self.current_store == store and self.conn is not self.master_conn:
            return
        db_path = self.store_db_path(store)
        key = str(db_path)
        if key not in self.store_conns:
            conn = sqlite3.connect(db_path, timeout=30)
            conn.row_factory = sqlite3.Row
            conn.execute("PRAGMA busy_timeout=30000")
            self.store_conns[key] = conn
        self.conn = self.store_conns[key]
        self.current_store = store
        if key not in self.initialized_store_dbs:
            if db_path.exists() and self._store_db_ready(self.conn):
                self.initialized_store_dbs.add(key)
                return
            try:
                self.init_db()
                self.conn.execute("""
                    INSERT OR IGNORE INTO stores (name, note, active, created_at)
                    VALUES (?, '', 1, ?)
                """, (store, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
                self.ensure_store_config(store)
                self.conn.commit()
                self.initialized_store_dbs.add(key)
            except sqlite3.OperationalError as exc:
                if "locked" not in str(exc).lower():
                    raise

    def import_excel(self, path: str, selected_store: str = "", report_type: str = "已结算", selected_month: str = "", progress_callback=None) -> ImportResult:
        """导入 Excel 流水，自动识别字段并写入当前店铺独立库。"""

        suffix = Path(path).suffix.lower()
        if suffix == ".csv":
            return self.import_csv(path, selected_store, report_type, selected_month, progress_callback)
        if load_workbook is None:
            raise RuntimeError("缺少 openpyxl，无法读取 Excel。请使用启动脚本附带的 Python 运行。")
        if selected_store:
            self.use_store(selected_store)
        report_type = report_type if report_type in REPORT_TYPES else "已结算"
        selected_month = normalize_month_token(selected_month)

        workbook = load_workbook(path, read_only=True, data_only=True)
        imported = 0
        skipped = 0
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cur = self.conn.cursor()
        merged_config_stores = set()

        for ws_index, ws in enumerate(workbook.worksheets, start=1):
            if progress_callback:
                progress_callback(f"读取工作表：{ws.title}", 0, 0)
            rows = ws.iter_rows(values_only=True)
            try:
                headers = [normalize_header(v) for v in next(rows)]
            except StopIteration:
                continue
            idx = {name: i for i, name in enumerate(headers)}
            if selected_store:
                if selected_store not in merged_config_stores:
                    self.merge_store_raw_columns(selected_store, headers)
                    merged_config_stores.add(selected_store)

            def cell(row, name):
                pos = idx.get(name)
                if pos is None or pos >= len(row):
                    return None
                return row[pos]

            def aliased(row, name):
                for candidate in HEADER_ALIASES.get(name, [name]):
                    value = cell(row, candidate)
                    if value not in (None, ""):
                        return value
                return None

            def amount_sum(row, name):
                total = Decimal("0.00")
                found = False
                for candidate in AMOUNT_ALIASES.get(name, [name]):
                    if candidate in idx:
                        total += money(cell(row, candidate))
                        found = True
                return total if found else money(cell(row, name))

            has_store = selected_store or any(name in idx for name in HEADER_ALIASES["店铺"])
            has_month_or_time = any(name in idx for name in HEADER_ALIASES["月份"] + HEADER_ALIASES["动账时间"])
            if not has_store or not has_month_or_time:
                missing = []
                if not has_store:
                    missing.append("店铺")
                if not has_month_or_time:
                    missing.append("月份或动账时间")
                raise ValueError(f"工作表 {ws.title} 无法自动识别关键列：{', '.join(missing)}")

            for row_index, row in enumerate(rows, start=2):
                def text(name):
                    return clean_cell_text(aliased(row, name))

                store = selected_store or text("店铺")
                if store and store not in merged_config_stores:
                    self.use_store(store)
                    self.merge_store_raw_columns(store, headers)
                    merged_config_stores.add(store)
                transaction_time = text("动账时间")
                month_label = text("月份")
                month_sort = selected_month or parse_month_sort(month_label, transaction_time)
                if selected_month:
                    month_label = selected_month
                elif not month_label:
                    month_label = month_sort
                flow_id = text("动账流水号")
                if not flow_id:
                    flow_id = f"{Path(path).name}:{ws.title}:{imported + skipped + 2}"
                if not store or not month_sort:
                    skipped += 1
                    continue
                raw_payload = {
                    headers[i]: json_default(row[i]) if i < len(row) else ""
                    for i in range(len(headers))
                    if headers[i]
                }
                values = (
                    Path(path).name, now, store, report_type, month_label, month_sort,
                    transaction_time, flow_id,
                    text("动账方向"), text("动账账户"),
                    db_float(aliased(row, "动账金额")), text("动账摘要"),
                    text("业务类型"), text("主订单编号"),
                    text("子订单编号"), text("售后单号"),
                    text("下单时间"), text("商品信息"),
                    text("商品编码"), text("售卖类型"),
                    db_float(amount_sum(row, "订单实付应结")), db_float(amount_sum(row, "平台补贴")),
                    db_float(amount_sum(row, "商家补贴")), db_float(amount_sum(row, "结算运费")),
                    db_float(amount_sum(row, "订单退款")), db_float(amount_sum(row, "佣金")),
                    db_float(amount_sum(row, "技术服务费")),
                    json.dumps(raw_payload, ensure_ascii=False, default=json_default),
                )
                try:
                    cur.execute("""
                        INSERT INTO transactions (
                            source_file, imported_at, store, report_type, month_label, month_sort,
                            transaction_time, flow_id, direction, account, amount,
                            summary, biz_type, main_order, sub_order, after_sale, order_time,
                            product_info, product_code, sale_type, paid_settlement,
                            platform_subsidy, merchant_subsidy, freight, refund,
                            commission, tech_fee, raw_payload
                        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """, values)
                    imported += 1
                except sqlite3.IntegrityError:
                    skipped += 1
                if progress_callback and row_index % 1000 == 0:
                    progress_callback(f"导入 {Path(path).name} / {ws.title}：已处理 {row_index - 1} 行", row_index - 1, 0)

        self.conn.commit()
        if progress_callback:
            progress_callback(f"完成：{Path(path).name}，新增 {imported} 条，跳过 {skipped} 条", imported + skipped, 0)
        return ImportResult(imported=imported, skipped=skipped, file_name=Path(path).name)

    def import_csv(self, path: str, selected_store: str = "", report_type: str = "已结算", selected_month: str = "", progress_callback=None) -> ImportResult:
        """导入 CSV 流水，兼容 UTF-8 BOM 和 GB18030 编码。"""

        if selected_store:
            self.use_store(selected_store)
        report_type = report_type if report_type in REPORT_TYPES else "已结算"
        selected_month = normalize_month_token(selected_month)
        imported = 0
        skipped = 0
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cur = self.conn.cursor()
        source_name = Path(path).name

        last_error = None
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                with open(path, "r", encoding=encoding, newline="") as fh:
                    reader = csv.reader(fh)
                    try:
                        headers = [normalize_header(v) for v in next(reader)]
                    except StopIteration:
                        return ImportResult(imported=0, skipped=0, file_name=source_name)
                    idx = {name: i for i, name in enumerate(headers)}
                    if selected_store:
                        self.merge_store_raw_columns(selected_store, headers)

                    def cell(row, name):
                        pos = idx.get(name)
                        if pos is None or pos >= len(row):
                            return None
                        return row[pos]

                    def aliased(row, name):
                        for candidate in HEADER_ALIASES.get(name, [name]):
                            value = cell(row, candidate)
                            if value not in (None, ""):
                                return value
                        return None

                    def amount_sum(row, name):
                        total = Decimal("0.00")
                        found = False
                        for candidate in AMOUNT_ALIASES.get(name, [name]):
                            if candidate in idx:
                                total += money(cell(row, candidate))
                                found = True
                        return total if found else money(cell(row, name))

                    has_store = selected_store or any(name in idx for name in HEADER_ALIASES["店铺"])
                    has_month_or_time = any(name in idx for name in HEADER_ALIASES["月份"] + HEADER_ALIASES["动账时间"])
                    if not has_store or not has_month_or_time:
                        missing = []
                        if not has_store:
                            missing.append("店铺")
                        if not has_month_or_time:
                            missing.append("月份或动账时间")
                        raise ValueError(f"CSV {source_name} 无法自动识别关键列：{', '.join(missing)}")

                    for row_index, row in enumerate(reader, start=2):
                        def text(name):
                            return clean_cell_text(aliased(row, name))

                        store = selected_store or text("店铺")
                        transaction_time = text("动账时间")
                        month_label = text("月份")
                        month_sort = selected_month or parse_month_sort(month_label, transaction_time)
                        if selected_month:
                            month_label = selected_month
                        elif not month_label:
                            month_label = month_sort
                        flow_id = text("动账流水号") or f"{source_name}:CSV:{row_index}"
                        if not store or not month_sort:
                            skipped += 1
                            continue
                        raw_payload = {
                            headers[i]: json_default(row[i]) if i < len(row) else ""
                            for i in range(len(headers))
                            if headers[i]
                        }
                        values = (
                            source_name, now, store, report_type, month_label, month_sort,
                            transaction_time, flow_id,
                            text("动账方向"), text("动账账户"),
                            db_float(aliased(row, "动账金额")), text("动账摘要"),
                            text("业务类型"), text("主订单编号"),
                            text("子订单编号"), text("售后单号"),
                            text("下单时间"), text("商品信息"),
                            text("商品编码"), text("售卖类型"),
                            db_float(amount_sum(row, "订单实付应结")), db_float(amount_sum(row, "平台补贴")),
                            db_float(amount_sum(row, "商家补贴")), db_float(amount_sum(row, "结算运费")),
                            db_float(amount_sum(row, "订单退款")), db_float(amount_sum(row, "佣金")),
                            db_float(amount_sum(row, "技术服务费")),
                            json.dumps(raw_payload, ensure_ascii=False, default=json_default),
                        )
                        try:
                            cur.execute("""
                                INSERT INTO transactions (
                                    source_file, imported_at, store, report_type, month_label, month_sort,
                                    transaction_time, flow_id, direction, account, amount,
                                    summary, biz_type, main_order, sub_order, after_sale, order_time,
                                    product_info, product_code, sale_type, paid_settlement,
                                    platform_subsidy, merchant_subsidy, freight, refund,
                                    commission, tech_fee, raw_payload
                                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                            """, values)
                            imported += 1
                        except sqlite3.IntegrityError:
                            skipped += 1
                        if progress_callback and row_index % 1000 == 0:
                            progress_callback(f"导入 {source_name}：已处理 {row_index - 1} 行", row_index - 1, 0)
                self.conn.commit()
                if progress_callback:
                    progress_callback(f"完成：{source_name}，新增 {imported} 条，跳过 {skipped} 条", imported + skipped, 0)
                return ImportResult(imported=imported, skipped=skipped, file_name=source_name)
            except UnicodeDecodeError as exc:
                self.conn.rollback()
                last_error = exc
                imported = 0
                skipped = 0
                continue
        raise RuntimeError(f"CSV 编码无法识别：{last_error}")

    def configured_stores(self, include_inactive=False):
        """读取已配置店铺，用于首页店铺下拉和店铺配置页。"""

        if include_inactive:
            rows = self.master_conn.execute("SELECT * FROM stores ORDER BY active DESC, name").fetchall()
        else:
            rows = self.master_conn.execute("SELECT * FROM stores WHERE active=1 ORDER BY name").fetchall()
        return rows

    def add_store(self, name, note=""):
        name = str(name or "").strip()
        if not name:
            raise ValueError("请输入店铺名称")
        self.master_conn.execute("""
            INSERT INTO stores (name, note, active, created_at, db_file)
            VALUES (?, ?, 1, ?, ?)
            ON CONFLICT(name) DO UPDATE SET note=excluded.note, active=1
        """, (name, note, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), self._default_store_db_file(name)))
        self.master_conn.commit()
        self.use_store(name)

    def delete_store(self, name):
        self.master_conn.execute("UPDATE stores SET active=0 WHERE name=?", (name,))
        self.master_conn.commit()

    def update_store(self, old_name, new_name, note="", active=1):
        old_name = str(old_name or "").strip()
        new_name = str(new_name or "").strip()
        if not old_name or not new_name:
            raise ValueError("店铺名称不能为空")
        if old_name != new_name:
            self.rename_store(old_name, new_name)
        self.master_conn.execute(
            "UPDATE stores SET note=?, active=? WHERE name=?",
            (note, 1 if int(active or 0) else 0, new_name),
        )
        self.master_conn.commit()

    def rename_store(self, old_name, new_name):
        old_name = str(old_name or "").strip()
        new_name = str(new_name or "").strip()
        if not old_name or not new_name:
            raise ValueError("店铺名称不能为空")
        if old_name == new_name:
            return
        if self._master_store_row(new_name):
            raise ValueError("新店铺名称已存在")
        row = self._master_store_row(old_name)
        if not row:
            raise ValueError("原店铺不存在")
        self.master_conn.execute("UPDATE stores SET name=? WHERE name=?", (new_name, old_name))
        self.master_conn.commit()
        self.use_store(new_name)
        self.conn.execute("DELETE FROM stores WHERE name=?", (old_name,))
        old_config = self.conn.execute("SELECT * FROM store_configs WHERE store=?", (old_name,)).fetchone()
        if old_config:
            self.conn.execute("""
                UPDATE store_configs
                SET raw_columns=?, summary_columns=?, frozen_columns=?, page_size=?, formula_note=?, updated_at=?
                WHERE store=?
            """, (
                old_config["raw_columns"], old_config["summary_columns"], old_config["frozen_columns"], old_config["page_size"], old_config["formula_note"],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"), new_name,
            ))
            self.conn.execute("DELETE FROM store_configs WHERE store=?", (old_name,))
        for table in ("transactions", "balances", "adjustments"):
            self.conn.execute(f"UPDATE {table} SET store=? WHERE store=?", (new_name, old_name))
        self.conn.commit()

    def stores(self):
        rows = self.master_conn.execute("SELECT name AS store FROM stores WHERE active=1 ORDER BY name").fetchall()
        return [r[0] for r in rows if r[0]]

    def ensure_store_config(self, store):
        if not store:
            return
        self.ensure_store_config_columns()
        self.conn.execute("""
            INSERT OR IGNORE INTO store_configs (store, raw_columns, summary_columns, frozen_columns, page_size, formula_note, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            store,
            json.dumps(RAW_COLUMNS, ensure_ascii=False),
            json.dumps(SUMMARY_COLUMNS + SUMMARY_EXTRA_COLUMNS, ensure_ascii=False),
            json.dumps(SUMMARY_COLUMNS[:3], ensure_ascii=False),
            1000,
            DEFAULT_FORMULA_NOTE,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ))

    def ensure_store_config_columns(self):
        cols = {row["name"] for row in self.conn.execute("PRAGMA table_info(store_configs)").fetchall()}
        if "frozen_columns" not in cols:
            self.conn.execute("ALTER TABLE store_configs ADD COLUMN frozen_columns TEXT")
        if "page_size" not in cols:
            self.conn.execute("ALTER TABLE store_configs ADD COLUMN page_size INTEGER DEFAULT 1000")

    def store_config(self, store):
        if store:
            self.use_store(store)
        row = self.conn.execute("SELECT * FROM store_configs WHERE store=?", (store,)).fetchone()
        if not row:
            self.ensure_store_config(store)
            row = self.conn.execute("SELECT * FROM store_configs WHERE store=?", (store,)).fetchone()
        if not row:
            return {
                "store": store,
                "raw_columns": RAW_COLUMNS,
                "summary_columns": SUMMARY_COLUMNS + SUMMARY_EXTRA_COLUMNS,
                "frozen_columns": SUMMARY_COLUMNS[:3],
                "page_size": 1000,
                "formula_note": DEFAULT_FORMULA_NOTE,
            }
        page_size = row["page_size"] if "page_size" in row.keys() else 1000
        return {
            "store": store,
            "raw_columns": self._loads_columns(row["raw_columns"], RAW_COLUMNS),
            "summary_columns": self._loads_columns(row["summary_columns"], SUMMARY_COLUMNS + SUMMARY_EXTRA_COLUMNS),
            "frozen_columns": self._loads_columns(row["frozen_columns"], SUMMARY_COLUMNS[:3]),
            "page_size": max(100, int(page_size or 1000)),
            "formula_note": row["formula_note"] or DEFAULT_FORMULA_NOTE,
        }

    def _loads_columns(self, text, fallback):
        try:
            values = json.loads(text or "[]")
            return [str(v).strip() for v in values if str(v).strip()] or list(fallback)
        except Exception:
            return list(fallback)

    def save_store_config(self, store, raw_columns, summary_columns, formula_note, frozen_columns=None, page_size=1000):
        """保存单店铺参数配置，包括字段、冻结栏、计算口径和分页行数。"""

        if store:
            self.use_store(store)
        self.ensure_store_config_columns()
        self.ensure_store_config(store)
        frozen_columns = frozen_columns or summary_columns[:2]
        page_size = max(100, int(page_size or 1000))
        self.conn.execute("""
            UPDATE store_configs
            SET raw_columns=?, summary_columns=?, frozen_columns=?, page_size=?, formula_note=?, updated_at=?
            WHERE store=?
        """, (
            json.dumps(raw_columns, ensure_ascii=False),
            json.dumps(summary_columns, ensure_ascii=False),
            json.dumps(frozen_columns, ensure_ascii=False),
            page_size,
            formula_note,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            store,
        ))
        self.conn.commit()

    def merge_store_raw_columns(self, store, headers):
        if not store:
            return
        self.use_store(store)
        config = self.store_config(store)
        columns = list(config["raw_columns"])
        changed = False
        for header in headers:
            header = str(header or "").strip()
            if header and header not in columns:
                columns.append(header)
                changed = True
        if changed:
            self.save_store_config(store, columns, config["summary_columns"], config["formula_note"], config.get("frozen_columns"), config.get("page_size", 1000))

    def months(self):
        if self.current_store:
            self.use_store(self.current_store)
        rows = self.conn.execute("SELECT DISTINCT month_sort, month_label FROM transactions UNION SELECT DISTINCT month_sort, month_label FROM balances ORDER BY month_sort").fetchall()
        return [{"month_sort": r["month_sort"], "month_label": r["month_label"]} for r in rows]

    def upsert_balance(self, store, month_label, month_sort, opening, account_ending, note, report_type="已结算"):
        self.use_store(store)
        report_type = report_type if report_type in REPORT_TYPES else "已结算"
        existing = self.conn.execute(
            "SELECT id FROM balances WHERE store=? AND report_type=? AND month_sort=?",
            (store, report_type, month_sort),
        ).fetchone()
        values = (month_label, db_float(opening), db_float(account_ending), note, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        if existing:
            self.conn.execute("""
                UPDATE balances
                SET month_label=?, opening_balance=?, account_ending_balance=?, note=?, updated_at=?
                WHERE id=?
            """, values + (existing["id"],))
        else:
            self.conn.execute("""
                INSERT INTO balances (store, report_type, month_label, month_sort, opening_balance, account_ending_balance, note, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (store, report_type, month_label, month_sort, db_float(opening), db_float(account_ending), note, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        self.conn.commit()

    def delete_month(self, store, month_sort, delete_balance=False, report_type=None):
        self.use_store(store)
        where_type = "" if not report_type else " AND report_type=?"
        params = [store, month_sort] + ([report_type] if report_type else [])
        self.conn.execute(f"DELETE FROM transactions WHERE store=? AND month_sort=?{where_type}", params)
        self.conn.execute(f"DELETE FROM adjustments WHERE store=? AND month_sort=?{where_type}", params)
        if delete_balance:
            self.conn.execute(f"DELETE FROM balances WHERE store=? AND month_sort=?{where_type}", params)
        self.conn.commit()

    def transaction_by_id(self, tx_id):
        return self.conn.execute("SELECT * FROM transactions WHERE id=?", (tx_id,)).fetchone()

    def update_transaction(self, tx_id, data):
        self.conn.execute("""
            UPDATE transactions SET
                store=?, month_label=?, month_sort=?, transaction_time=?,
                direction=?, account=?, amount=?, summary=?, biz_type=?
            WHERE id=?
        """, (
            data["店铺"], data["月份显示"], data["月份排序(YYYY-MM)"], data["动账时间"],
            data["动账方向"], data["动账账户"], db_float(data["动账金额"]),
            data["动账摘要"], data["业务类型"], tx_id,
        ))
        self.conn.commit()

    def delete_transaction(self, tx_id):
        self.conn.execute("DELETE FROM transactions WHERE id=?", (tx_id,))
        self.conn.commit()

    def add_adjustment(self, store, month_label, month_sort, target_column, item, amount, note, report_type="已结算"):
        self.use_store(store)
        report_type = report_type if report_type in REPORT_TYPES else "已结算"
        self.conn.execute("""
            INSERT INTO adjustments (store, report_type, month_label, month_sort, target_column, item, amount, note, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (store, report_type, month_label, month_sort, target_column, item, db_float(amount), note, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        self.conn.commit()

    def delete_adjustment(self, adj_id):
        self.conn.execute("DELETE FROM adjustments WHERE id=?", (adj_id,))
        self.conn.commit()

    def monthly_summaries(self, store_filter="", month_filter="", aggregate=True, report_type_filter=""):
        """按店铺、报表类型、年月区间生成汇总行和区间合计行。"""

        store_filter = str(store_filter or "").strip()
        if not store_filter:
            return []
        self.use_store(store_filter)
        months = parse_month_filter(month_filter)
        tx_rows = self._monthly_base_rows(store_filter, months, report_type_filter)
        month_rows = [self._build_month_summary(row) for row in tx_rows]
        month_rows.extend(self._balance_only_rows(store_filter, months, report_type_filter))
        month_rows = sorted(month_rows, key=lambda x: (x["store"], x["report_type"], x["month_sort"]))
        if not aggregate or len(months) <= 1:
            return month_rows
        return self._aggregate_rows(month_rows, months)

    def _month_where(self, where, params, months):
        if months:
            where.append("month_sort IN (%s)" % ",".join("?" for _ in months))
            params.extend(months)

    def _monthly_base_rows(self, store_filter="", months=None, report_type_filter=""):
        months = months or []
        params = []
        where = []
        if store_filter:
            where.append("store LIKE ?")
            params.append(f"%{store_filter}%")
        if report_type_filter:
            where.append("report_type=?")
            params.append(report_type_filter)
        self._month_where(where, params, months)
        where_sql = "WHERE " + " AND ".join(where) if where else ""
        return self.conn.execute(f"""
            SELECT store, report_type, month_label, month_sort,
                   SUM(CASE WHEN summary='账户余额提现' THEN amount ELSE 0 END) AS withdraw_amount,
                   SUM(paid_settlement) AS paid_settlement,
                   SUM(platform_subsidy) AS platform_subsidy,
                   SUM(merchant_subsidy) AS merchant_subsidy,
                   SUM(freight) AS freight,
                   SUM(refund) AS refund,
                   SUM(paid_settlement + platform_subsidy + merchant_subsidy + freight + refund) AS income_total,
                   SUM(commission) AS commission,
                   SUM(tech_fee) AS tech_fee,
                   SUM(commission + tech_fee) AS expense_amount,
                   SUM(paid_settlement + platform_subsidy + merchant_subsidy + freight + refund + commission + tech_fee) AS settlement_amount
            FROM transactions
            {where_sql}
            GROUP BY store, report_type, month_sort
        """, params).fetchall()

    def _build_month_summary(self, row):
        """把数据库聚合结果、余额和调整记录组装成单月汇总。"""

        report_type = row["report_type"] or "已结算"
        balance = self.balance_for(row["store"], row["month_sort"], report_type)
        adjustments_by_target = self.adjustment_totals_by_target(row["store"], row["month_sort"], report_type)
        paid_settlement = money(row["paid_settlement"]) + adjustments_by_target.get("订单实付应结", Decimal("0.00"))
        platform_subsidy = money(row["platform_subsidy"]) + adjustments_by_target.get("平台补贴", Decimal("0.00"))
        merchant_subsidy = money(row["merchant_subsidy"]) + adjustments_by_target.get("商家补贴", Decimal("0.00"))
        freight = money(row["freight"]) + adjustments_by_target.get("结算运费", Decimal("0.00"))
        refund = money(row["refund"]) + adjustments_by_target.get("订单退款", Decimal("0.00"))
        commission = money(row["commission"]) + adjustments_by_target.get("已结算佣金", Decimal("0.00"))
        tech_fee = money(row["tech_fee"]) + adjustments_by_target.get("技术服务费", Decimal("0.00"))
        withdraw_amount = money(row["withdraw_amount"]) + adjustments_by_target.get("提现金额", Decimal("0.00"))
        opening = money(balance["opening_balance"] if balance else 0) + adjustments_by_target.get("期初金额", Decimal("0.00"))
        income_total = paid_settlement + platform_subsidy + merchant_subsidy + freight + refund
        expense_amount = commission + tech_fee
        settlement_amount = income_total + expense_amount
        ending = opening + income_total + expense_amount + withdraw_amount
        account_ending = None if not balance or balance["account_ending_balance"] is None else money(balance["account_ending_balance"])
        diff = None if account_ending is None else ending - account_ending
        summary = {
            "store": row["store"], "report_type": report_type, "month_label": row["month_label"], "month_sort": row["month_sort"],
            "month_sorts": [row["month_sort"]], "is_aggregate": False,
            "paid_settlement": paid_settlement, "platform_subsidy": platform_subsidy,
            "merchant_subsidy": merchant_subsidy, "freight": freight, "refund": refund,
            "income_total": income_total, "commission": commission, "tech_fee": tech_fee,
            "expense_amount": expense_amount, "settlement_amount": settlement_amount,
            "withdraw_amount": withdraw_amount, "opening_balance": opening,
            "adjustments": self.adjustment_total(row["store"], row["month_sort"], report_type),
            "adjustments_by_target": adjustments_by_target,
            "adjustment_notes": self.adjustment_notes(row["store"], [row["month_sort"]], report_type),
            "ending_balance": ending, "account_ending": account_ending, "difference": diff,
        }
        self.apply_store_formulas(row["store"], summary)
        summary["difference"] = None if summary["account_ending"] is None else summary["ending_balance"] - summary["account_ending"]
        return summary

    def _balance_only_rows(self, store_filter="", months=None, report_type_filter=""):
        months = months or []
        result = []
        rows = self.conn.execute("""
            SELECT b.* FROM balances b
            LEFT JOIN transactions t ON t.store=b.store AND t.report_type=b.report_type AND t.month_sort=b.month_sort
            WHERE t.id IS NULL
            ORDER BY b.store, b.report_type, b.month_sort
        """).fetchall()
        for row in rows:
            if store_filter and store_filter not in row["store"]:
                continue
            if report_type_filter and report_type_filter != row["report_type"]:
                continue
            if months and row["month_sort"] not in months:
                continue
            opening = money(row["opening_balance"])
            report_type = row["report_type"] or "已结算"
            adjustments_by_target = self.adjustment_totals_by_target(row["store"], row["month_sort"], report_type)
            paid_settlement = adjustments_by_target.get("订单实付应结", Decimal("0.00"))
            platform_subsidy = adjustments_by_target.get("平台补贴", Decimal("0.00"))
            merchant_subsidy = adjustments_by_target.get("商家补贴", Decimal("0.00"))
            freight = adjustments_by_target.get("结算运费", Decimal("0.00"))
            refund = adjustments_by_target.get("订单退款", Decimal("0.00"))
            commission = adjustments_by_target.get("已结算佣金", Decimal("0.00"))
            tech_fee = adjustments_by_target.get("技术服务费", Decimal("0.00"))
            withdraw_amount = adjustments_by_target.get("提现金额", Decimal("0.00"))
            opening = opening + adjustments_by_target.get("期初金额", Decimal("0.00"))
            income_total = paid_settlement + platform_subsidy + merchant_subsidy + freight + refund
            expense_amount = commission + tech_fee
            settlement_amount = income_total + expense_amount
            ending = opening + income_total + expense_amount + withdraw_amount
            account_ending = None if row["account_ending_balance"] is None else money(row["account_ending_balance"])
            summary = {
                "store": row["store"], "report_type": report_type, "month_label": row["month_label"], "month_sort": row["month_sort"],
                "month_sorts": [row["month_sort"]], "is_aggregate": False,
                "paid_settlement": paid_settlement, "platform_subsidy": platform_subsidy, "merchant_subsidy": merchant_subsidy,
                "freight": freight, "refund": refund, "income_total": income_total,
                "commission": commission, "tech_fee": tech_fee, "expense_amount": expense_amount,
                "settlement_amount": settlement_amount, "withdraw_amount": withdraw_amount,
                "opening_balance": opening, "adjustments": self.adjustment_total(row["store"], row["month_sort"], report_type),
                "adjustments_by_target": adjustments_by_target,
                "adjustment_notes": self.adjustment_notes(row["store"], [row["month_sort"]], report_type),
                "ending_balance": ending, "account_ending": account_ending,
                "difference": None if account_ending is None else ending - account_ending,
            }
            self.apply_store_formulas(row["store"], summary)
            summary["difference"] = None if summary["account_ending"] is None else summary["ending_balance"] - summary["account_ending"]
            result.append(summary)
        return result

    def _aggregate_rows(self, rows, months):
        """把多个月汇总为区间合计，期初取首月、期末取末月。"""

        result = []
        by_store = {}
        for row in rows:
            by_store.setdefault((row["store"], row["report_type"]), []).append(row)
        numeric_sum_fields = [
            "paid_settlement", "platform_subsidy", "merchant_subsidy", "freight", "refund",
            "income_total", "commission", "tech_fee", "expense_amount", "settlement_amount",
            "withdraw_amount", "adjustments",
        ]
        for (store, report_type), store_rows in by_store.items():
            store_rows = sorted(store_rows, key=lambda r: r["month_sort"])
            first_month = store_rows[0]["month_sort"]
            last_month = store_rows[-1]["month_sort"]
            first_balance = self.balance_for(store, first_month, report_type)
            last_balance = self.balance_for(store, last_month, report_type)
            aggregate = {
                "store": store,
                "report_type": report_type,
                "month_label": f"{months[0]}至{months[-1]}",
                "month_sort": f"{months[0]}~{months[-1]}",
                "month_sorts": [r["month_sort"] for r in store_rows],
                "is_aggregate": True,
                "opening_balance": money(first_balance["opening_balance"] if first_balance else store_rows[0]["opening_balance"]),
                "account_ending": None if not last_balance or last_balance["account_ending_balance"] is None else money(last_balance["account_ending_balance"]),
                "adjustment_notes": self.adjustment_notes(store, [r["month_sort"] for r in store_rows], report_type),
            }
            for field in numeric_sum_fields:
                aggregate[field] = sum((money(r[field]) for r in store_rows), Decimal("0.00"))
            aggregate["ending_balance"] = (
                aggregate["opening_balance"] + aggregate["income_total"] +
                aggregate["expense_amount"] + aggregate["withdraw_amount"]
            )
            self.apply_store_formulas(store, aggregate)
            aggregate["difference"] = None if aggregate["account_ending"] is None else aggregate["ending_balance"] - aggregate["account_ending"]
            result.append(aggregate)
            result.extend(store_rows)
        return result

    def apply_store_formulas(self, store, summary):
        """应用店铺自定义计算口径，没有配置时使用默认公式。"""

        formulas = parse_formula_lines(self.store_config(store)["formula_note"])
        for target, expr in formulas:
            key = FIELD_KEYS[target]
            try:
                summary[key] = eval_money_formula(expr, summary)
            except Exception:
                continue

    def balance_for(self, store, month_sort, report_type="已结算"):
        self.use_store(store)
        return self.conn.execute(
            "SELECT * FROM balances WHERE store=? AND report_type=? AND month_sort=?",
            (store, report_type, month_sort),
        ).fetchone()

    def adjustment_total(self, store, month_sort, report_type="已结算") -> Decimal:
        self.use_store(store)
        row = self.conn.execute(
            "SELECT SUM(amount) AS total FROM adjustments WHERE store=? AND report_type=? AND month_sort=?",
            (store, report_type, month_sort),
        ).fetchone()
        return money(row["total"] if row else 0)

    def adjustment_totals_by_target(self, store, month_sort, report_type="已结算"):
        self.use_store(store)
        rows = self.conn.execute("""
            SELECT COALESCE(target_column, '备查（不影响汇总）') AS target_column, SUM(amount) AS total
            FROM adjustments
            WHERE store=? AND report_type=? AND month_sort=?
            GROUP BY COALESCE(target_column, '备查（不影响汇总）')
        """, (store, report_type, month_sort)).fetchall()
        return {row["target_column"]: money(row["total"]) for row in rows}

    def adjustment_notes(self, store, months, report_type="已结算"):
        if not months:
            return ""
        self.use_store(store)
        rows = self.conn.execute(f"""
            SELECT month_sort, target_column, item, amount, note
            FROM adjustments
            WHERE store=? AND report_type=? AND month_sort IN ({",".join("?" for _ in months)})
            ORDER BY month_sort, id
        """, [store, report_type] + list(months)).fetchall()
        return "；".join(
            f"{r['month_sort']} [{r['target_column'] or '备查'}] {r['item']} {money_text(r['amount'])} {r['note'] or ''}".strip()
            for r in rows
        )

    def _detail_column_map(self):
        return {
            "ID": "id", "店铺": "store", "报表类型": "report_type", "月份": "month_sort", "年月": "month_sort", "动账时间": "transaction_time",
            "动账流水号": "flow_id", "动账方向": "direction", "动账账户": "account",
            "动账金额": "amount", "动账摘要": "summary", "业务类型": "biz_type",
            "主订单编号": "main_order", "子订单编号": "sub_order", "售后单号": "after_sale",
            "下单时间": "order_time", "商品信息": "product_info", "商品编码": "product_code",
            "售卖类型": "sale_type", "订单实付应结": "paid_settlement", "平台补贴": "platform_subsidy",
            "商家补贴": "merchant_subsidy", "结算运费": "freight", "订单退款": "refund",
            "佣金": "commission", "技术服务费": "tech_fee",
        }

    def _detail_filter_sql(self, filters, params):
        """把明细筛选条件转换为 SQL where 片段和参数。"""

        sql = ""
        column_map = self._detail_column_map()
        for column, value in filters or []:
            value = str(value or "").strip()
            filter_col = column_map.get(column)
            if filter_col and value:
                sql += f" AND {clean_sql_expr(filter_col)}=?"
                params.append(clean_cell_text(value))
            elif column and value:
                raw_path = '$."' + str(column).replace('"', '\\"') + '"'
                raw_path_sql = sql_literal(raw_path)
                sql += f" AND COALESCE({clean_sql_expr(f'json_extract(raw_payload, {raw_path_sql})')}, '')=?"
                params.append(clean_cell_text(value))
        return sql

    def details(self, store, month_sort=None, months=None, sort_order="ASC", sort_column="动账时间", report_type_filter="", limit=None, offset=0, filter_column="", filter_value="", filters=None):
        """分页读取明细流水，排序和筛选尽量下推到 SQLite。"""

        months = months or ([month_sort] if month_sort else [])
        if not store or not months:
            return []
        self.use_store(store)
        order = "DESC" if str(sort_order).upper() == "DESC" else "ASC"
        sort_map = self._detail_column_map()
        order_col = sort_map.get(sort_column, "transaction_time")
        type_sql = " AND report_type=?" if report_type_filter else ""
        params = [store] + list(months) + ([report_type_filter] if report_type_filter else [])
        merged_filters = list(filters or [])
        if filter_column and filter_value:
            merged_filters.append((filter_column, filter_value))
        filter_sql = self._detail_filter_sql(merged_filters, params)
        limit_sql = ""
        if limit:
            limit_sql = " LIMIT ? OFFSET ?"
            params.extend([int(limit), int(offset or 0)])
        return self.conn.execute(f"""
            SELECT * FROM transactions WHERE store=? AND month_sort IN ({",".join("?" for _ in months)}){type_sql}{filter_sql}
            ORDER BY {order_col} {order}, id {order}
            {limit_sql}
        """, params).fetchall()

    def details_count(self, store, month_sort=None, months=None, report_type_filter="", filter_column="", filter_value="", filters=None):
        months = months or ([month_sort] if month_sort else [])
        if not store or not months:
            return 0
        self.use_store(store)
        type_sql = " AND report_type=?" if report_type_filter else ""
        params = [store] + list(months) + ([report_type_filter] if report_type_filter else [])
        merged_filters = list(filters or [])
        if filter_column and filter_value:
            merged_filters.append((filter_column, filter_value))
        filter_sql = self._detail_filter_sql(merged_filters, params)
        row = self.conn.execute(f"""
            SELECT COUNT(*) AS count
            FROM transactions
            WHERE store=? AND month_sort IN ({",".join("?" for _ in months)}){type_sql}{filter_sql}
        """, params).fetchone()
        return int(row["count"] if row else 0)

    def detail_distinct_values(self, store, months, report_type_filter, column, filters=None, limit=500):
        """读取筛选下拉候选值，用于 Excel 风格的精确筛选体验。"""

        if not store or not months:
            return []
        self.use_store(store)
        column_map = self._detail_column_map()
        db_col = column_map.get(column)
        if db_col:
            value_expr = clean_sql_expr(db_col)
        else:
            raw_path = '$."' + str(column).replace('"', '\\"') + '"'
            raw_path_sql = sql_literal(raw_path)
            value_expr = clean_sql_expr(f"json_extract(raw_payload, {raw_path_sql})")
        params = [store] + list(months)
        type_sql = " AND report_type=?" if report_type_filter else ""
        if report_type_filter:
            params.append(report_type_filter)
        filter_sql = self._detail_filter_sql([(c, v) for c, v in (filters or []) if c != column], params)
        params.append(int(limit))
        rows = self.conn.execute(f"""
            SELECT DISTINCT {value_expr} AS value
            FROM transactions
            WHERE store=? AND month_sort IN ({",".join("?" for _ in months)}){type_sql}{filter_sql}
                  AND {value_expr} IS NOT NULL AND TRIM(CAST({value_expr} AS TEXT)) <> ''
            ORDER BY value
            LIMIT ?
        """, params).fetchall()
        return [str(row["value"]) for row in rows]

    def raw_rows_for_export(self, store_filter="", month_filter="", report_type_filter=""):
        store_filter = str(store_filter or "").strip()
        if not store_filter:
            return [], []
        self.use_store(store_filter)
        months = parse_month_filter(month_filter)
        params = []
        where = []
        if store_filter:
            where.append("store LIKE ?")
            params.append(f"%{store_filter}%")
        if report_type_filter:
            where.append("report_type=?")
            params.append(report_type_filter)
        self._month_where(where, params, months)
        where_sql = "WHERE " + " AND ".join(where) if where else ""
        rows = self.conn.execute(f"""
            SELECT * FROM transactions {where_sql}
            ORDER BY store, report_type, month_sort, transaction_time, id
        """, params).fetchall()
        columns = []
        payloads = []
        for row in rows:
            try:
                payload = json.loads(row["raw_payload"] or "{}")
            except Exception:
                payload = {}
            if not payload:
                payload = {col: row[col] if col in row.keys() else "" for col in RAW_COLUMNS}
            for key in payload.keys():
                if key not in columns:
                    columns.append(key)
            payloads.append((row, payload))
        return columns, payloads

    def raw_rows_count_for_export(self, store_filter="", month_filter="", report_type_filter=""):
        store_filter = str(store_filter or "").strip()
        if not store_filter:
            return 0
        self.use_store(store_filter)
        months = parse_month_filter(month_filter)
        params = []
        where = []
        if store_filter:
            where.append("store LIKE ?")
            params.append(f"%{store_filter}%")
        if report_type_filter:
            where.append("report_type=?")
            params.append(report_type_filter)
        self._month_where(where, params, months)
        where_sql = "WHERE " + " AND ".join(where) if where else ""
        row = self.conn.execute(f"SELECT COUNT(*) AS count FROM transactions {where_sql}", params).fetchone()
        return int(row["count"] if row else 0)

    def raw_rows_for_export_page(self, store_filter="", month_filter="", report_type_filter="", limit=5000, offset=0):
        """导出原始表格时按批读取，避免一次性加载大表。"""

        store_filter = str(store_filter or "").strip()
        if not store_filter:
            return []
        self.use_store(store_filter)
        months = parse_month_filter(month_filter)
        params = []
        where = []
        if store_filter:
            where.append("store LIKE ?")
            params.append(f"%{store_filter}%")
        if report_type_filter:
            where.append("report_type=?")
            params.append(report_type_filter)
        self._month_where(where, params, months)
        where_sql = "WHERE " + " AND ".join(where) if where else ""
        params.extend([int(limit), int(offset)])
        rows = self.conn.execute(f"""
            SELECT * FROM transactions {where_sql}
            ORDER BY store, report_type, month_sort, transaction_time, id
            LIMIT ? OFFSET ?
        """, params).fetchall()
        payloads = []
        for row in rows:
            try:
                payload = json.loads(row["raw_payload"] or "{}")
            except Exception:
                payload = {}
            payloads.append((row, payload))
        return payloads

    def difference_groups(self, store, month_sort=None, months=None, report_type="已结算"):
        """按动账方向和摘要分组差异相关流水，供差异明细页展示。"""

        months = months or ([month_sort] if month_sort else [])
        if not store or not months:
            return [], []
        self.use_store(store)
        rows = self.conn.execute("""
            SELECT direction, summary, COUNT(*) AS count, SUM(amount) AS amount,
                   SUM(paid_settlement) AS paid_settlement,
                   SUM(platform_subsidy) AS platform_subsidy,
                   SUM(merchant_subsidy) AS merchant_subsidy,
                   SUM(freight) AS freight,
                   SUM(refund) AS refund,
                   SUM(paid_settlement + platform_subsidy + merchant_subsidy + freight + refund) AS income_total,
                   SUM(commission) AS commission,
                   SUM(tech_fee) AS tech_fee,
                   SUM(commission + tech_fee) AS expense_amount
            FROM transactions
            WHERE store=? AND report_type=? AND month_sort IN (%s)
            GROUP BY direction, summary
            ORDER BY direction, summary
        """ % ",".join("?" for _ in months), [store, report_type] + list(months)).fetchall()
        adjs = self.conn.execute(f"""
            SELECT * FROM adjustments WHERE store=? AND report_type=? AND month_sort IN ({",".join("?" for _ in months)})
            ORDER BY month_sort, id
        """, [store, report_type] + list(months)).fetchall()
        return rows, adjs


class BalanceDialog(simpledialog.Dialog):
    def __init__(self, parent, title, initial=None):
        self.initial = initial or {}
        self.result = None
        super().__init__(parent, title)

    def body(self, master):
        labels = ["店铺", "月份显示", "月份排序(YYYY-MM)", "期初金额", "店铺期末余额", "备注"]
        self.entries = {}
        for i, label in enumerate(labels):
            ttk.Label(master, text=label).grid(row=i, column=0, sticky="e", padx=8, pady=6)
            entry = ttk.Entry(master, width=34)
            entry.grid(row=i, column=1, sticky="ew", padx=8, pady=6)
            self.entries[label] = entry
        values = {
            "店铺": self.initial.get("store", ""),
            "月份显示": self.initial.get("month_label", ""),
            "月份排序(YYYY-MM)": self.initial.get("month_sort", ""),
            "期初金额": self.initial.get("opening_balance", "0"),
            "店铺期末余额": self.initial.get("account_ending", "0"),
            "备注": self.initial.get("note", ""),
        }
        for label, value in values.items():
            self.entries[label].insert(0, "" if value is None else str(value))
        return self.entries["店铺"]

    def validate(self):
        try:
            if not self.entries["店铺"].get().strip():
                raise ValueError("请输入店铺")
            if not self.entries["月份排序(YYYY-MM)"].get().strip():
                raise ValueError("请输入月份排序，例如 2026-06")
            money(self.entries["期初金额"].get())
            money(self.entries["店铺期末余额"].get())
            return True
        except Exception as exc:
            messagebox.showerror("输入有误", str(exc), parent=self)
            return False

    def apply(self):
        self.result = {
            "store": self.entries["店铺"].get().strip(),
            "month_label": self.entries["月份显示"].get().strip() or self.entries["月份排序(YYYY-MM)"].get().strip(),
            "month_sort": self.entries["月份排序(YYYY-MM)"].get().strip(),
            "opening_balance": self.entries["期初金额"].get().strip(),
            "account_ending": self.entries["店铺期末余额"].get().strip(),
            "note": self.entries["备注"].get().strip(),
        }


class AdjustmentDialog(simpledialog.Dialog):
    def __init__(self, parent, title, initial):
        self.initial = initial
        self.result = None
        super().__init__(parent, title)

    def body(self, master):
        ttk.Label(master, text="调整列").grid(row=0, column=0, sticky="e", padx=8, pady=6)
        ttk.Label(master, text="调整事项").grid(row=1, column=0, sticky="e", padx=8, pady=6)
        ttk.Label(master, text="金额").grid(row=2, column=0, sticky="e", padx=8, pady=6)
        ttk.Label(master, text="说明").grid(row=3, column=0, sticky="e", padx=8, pady=6)
        self.target_var = tk.StringVar(value=ADJUSTABLE_COLUMNS[0])
        self.target = ttk.Combobox(master, textvariable=self.target_var, values=ADJUSTABLE_COLUMNS, width=32, state="readonly")
        self.item = ttk.Entry(master, width=34)
        self.amount = ttk.Entry(master, width=34)
        self.note = ttk.Entry(master, width=34)
        self.target.grid(row=0, column=1, padx=8, pady=6)
        self.item.grid(row=1, column=1, padx=8, pady=6)
        self.amount.grid(row=2, column=1, padx=8, pady=6)
        self.note.grid(row=3, column=1, padx=8, pady=6)
        self.item.insert(0, "手工差异调整")
        self.amount.insert(0, "0")
        return self.target

    def validate(self):
        try:
            if not self.item.get().strip():
                raise ValueError("请输入调整事项")
            money(self.amount.get())
            return True
        except Exception as exc:
            messagebox.showerror("输入有误", str(exc), parent=self)
            return False

    def apply(self):
        self.result = {
            "target_column": self.target_var.get().strip(),
            "item": self.item.get().strip(),
            "amount": self.amount.get().strip(),
            "note": self.note.get().strip(),
        }


class TransactionDialog(simpledialog.Dialog):
    def __init__(self, parent, title, initial):
        self.initial = initial
        self.result = None
        super().__init__(parent, title)

    def body(self, master):
        fields = ["店铺", "月份显示", "月份排序(YYYY-MM)", "动账时间", "动账方向", "动账账户", "动账金额", "动账摘要", "业务类型"]
        self.entries = {}
        for i, field in enumerate(fields):
            ttk.Label(master, text=field).grid(row=i, column=0, sticky="e", padx=8, pady=5)
            entry = ttk.Entry(master, width=42)
            entry.grid(row=i, column=1, sticky="ew", padx=8, pady=5)
            entry.insert(0, str(self.initial.get(field, "") or ""))
            self.entries[field] = entry
        return self.entries["动账金额"]

    def validate(self):
        try:
            if not self.entries["店铺"].get().strip():
                raise ValueError("请输入店铺")
            if not self.entries["月份排序(YYYY-MM)"].get().strip():
                raise ValueError("请输入月份排序")
            money(self.entries["动账金额"].get())
            return True
        except Exception as exc:
            messagebox.showerror("输入有误", str(exc), parent=self)
            return False

    def apply(self):
        self.result = {key: entry.get().strip() for key, entry in self.entries.items()}


class ImportStoreDialog(simpledialog.Dialog):
    def __init__(self, parent, stores):
        self.stores = stores
        self.result = None
        super().__init__(parent, "选择导入店铺")

    def body(self, master):
        ttk.Label(master, text="导入归属店铺").grid(row=0, column=0, sticky="e", padx=8, pady=8)
        self.store_var = tk.StringVar()
        self.store_box = ttk.Combobox(master, textvariable=self.store_var, values=self.stores, width=34, state="readonly")
        self.store_box.grid(row=0, column=1, sticky="ew", padx=8, pady=8)
        if self.stores:
            self.store_box.current(0)
        ttk.Label(master, text="说明：选择后，本次导入的流水全部按该店铺处理；Excel 没有“店铺”列也可以导入。").grid(
            row=1, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 8)
        )
        return self.store_box

    def validate(self):
        if not self.store_var.get().strip():
            messagebox.showerror("请选择店铺", "请先选择一个已配置店铺。", parent=self)
            return False
        return True

    def apply(self):
        self.result = self.store_var.get().strip()


class StoreManageDialog(tk.Toplevel):
    def __init__(self, parent, repo, on_changed):
        super().__init__(parent)
        self.repo = repo
        self.on_changed = on_changed
        self.title("店铺配置")
        self.geometry("520x420")
        self.transient(parent)
        self.create_widgets()
        self.refresh()

    def create_widgets(self):
        form = ttk.Frame(self, padding=10)
        form.pack(fill="x")
        ttk.Label(form, text="店铺名称").grid(row=0, column=0, padx=4, pady=6, sticky="e")
        self.name_entry = ttk.Entry(form, width=28)
        self.name_entry.grid(row=0, column=1, padx=4, pady=6, sticky="ew")
        ttk.Label(form, text="备注").grid(row=1, column=0, padx=4, pady=6, sticky="e")
        self.note_entry = ttk.Entry(form, width=28)
        self.note_entry.grid(row=1, column=1, padx=4, pady=6, sticky="ew")
        ttk.Button(form, text="新增/启用店铺", command=self.add_store).grid(row=0, column=2, rowspan=2, padx=8, pady=6)
        form.columnconfigure(1, weight=1)

        self.tree = ttk.Treeview(self, columns=["店铺", "备注", "状态"], show="headings")
        self.tree.pack(fill="both", expand=True, padx=10, pady=(0, 8))
        for col in ["店铺", "备注", "状态"]:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=150, anchor="center")

        actions = ttk.Frame(self, padding=(10, 0, 10, 10))
        actions.pack(fill="x")
        ttk.Button(actions, text="停用选中店铺", command=self.delete_store).pack(side="left", padx=4)
        ttk.Button(actions, text="关闭", command=self.destroy).pack(side="right", padx=4)

    def refresh(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for row in self.repo.configured_stores(include_inactive=True):
            status = "启用" if row["active"] else "停用"
            self.tree.insert("", "end", values=[row["name"], row["note"] or "", status], iid=row["name"])

    def add_store(self):
        try:
            self.repo.add_store(self.name_entry.get(), self.note_entry.get())
            self.name_entry.delete(0, "end")
            self.note_entry.delete(0, "end")
            self.refresh()
            self.on_changed()
        except Exception as exc:
            messagebox.showerror("保存失败", str(exc), parent=self)

    def delete_store(self):
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("未选择", "请先选择一个店铺。", parent=self)
            return
        name = selection[0]
        if not messagebox.askyesno("确认停用", f"确定停用店铺“{name}”吗？历史流水不会删除。", parent=self):
            return
        self.repo.delete_store(name)
        self.refresh()
        self.on_changed()


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1280x760")
        self.minsize(1080, 680)
        self.repo = Repository(DB_PATH)
        self.selected_summary = None
        self.create_widgets()
        self.refresh_all()

    def create_widgets(self):
        style = ttk.Style(self)
        if "clam" in style.theme_names():
            style.theme_use("clam")
        self.configure(bg="#f5f7fb")
        style.configure(".", font=("Microsoft YaHei UI", 9))
        style.configure("TFrame", background="#f5f7fb")
        style.configure("TLabel", background="#f5f7fb", foreground="#243042")
        style.configure("App.TFrame", background="#f5f7fb")
        style.configure("Header.TFrame", background="#17233c")
        style.configure("Surface.TLabelframe", background="#f5f7fb", bordercolor="#d8dee9", relief="solid")
        style.configure("Surface.TLabelframe.Label", background="#f5f7fb", foreground="#283447", font=("Microsoft YaHei UI", 10, "bold"))
        style.configure("Treeview", rowheight=28, background="#ffffff", fieldbackground="#ffffff", foreground="#243042", bordercolor="#d8dee9")
        style.configure("Treeview.Heading", background="#eef2f7", foreground="#1f2a3d", font=("Microsoft YaHei UI", 9, "bold"), relief="flat")
        style.map("Treeview", background=[("selected", "#dbeafe")], foreground=[("selected", "#0f172a")])
        style.configure("Title.TLabel", font=("Microsoft YaHei UI", 15, "bold"), background="#17233c", foreground="#ffffff")
        style.configure("Subtitle.TLabel", font=("Microsoft YaHei UI", 9), background="#17233c", foreground="#cbd5e1")
        style.configure("Muted.TLabel", background="#f5f7fb", foreground="#64748b")
        style.configure("Primary.TButton", font=("Microsoft YaHei UI", 9, "bold"), padding=(12, 7), background="#2563eb", foreground="#ffffff")
        style.map("Primary.TButton", background=[("active", "#1d4ed8")], foreground=[("disabled", "#dbeafe")])
        style.configure("Action.TButton", padding=(10, 6), background="#ffffff", foreground="#1f2a3d")
        style.configure("Danger.TButton", padding=(10, 6), background="#fff1f2", foreground="#be123c")
        style.configure("TEntry", padding=(4, 3), fieldbackground="#ffffff")

        top = ttk.Frame(self, padding=(16, 12), style="Header.TFrame")
        top.pack(fill="x")
        title_box = ttk.Frame(top, style="Header.TFrame")
        title_box.pack(side="left", fill="x", expand=True)
        ttk.Label(title_box, text=APP_TITLE, style="Title.TLabel").pack(anchor="w")
        ttk.Label(title_box, text="第三方支付流水导入、月度汇总、余额核对与差异追踪", style="Subtitle.TLabel").pack(anchor="w", pady=(2, 0))
        ttk.Button(top, text="导入Excel流水", command=self.import_excel, style="Primary.TButton").pack(side="right", padx=(8, 0))
        ttk.Button(top, text="店铺配置", command=self.manage_stores, style="Action.TButton").pack(side="right", padx=4)
        ttk.Button(top, text="录入/修改余额", command=self.edit_balance, style="Action.TButton").pack(side="right", padx=4)
        ttk.Button(top, text="导出汇总CSV", command=self.export_summary, style="Action.TButton").pack(side="right", padx=4)

        filters = ttk.LabelFrame(self, text="查询条件", style="Surface.TLabelframe")
        filters.pack(fill="x", padx=14, pady=(12, 8))
        ttk.Label(filters, text="店铺").pack(side="left", padx=(10, 4), pady=8)
        self.store_filter = ttk.Entry(filters, width=20)
        self.store_filter.pack(side="left", padx=4)
        ttk.Label(filters, text="月份").pack(side="left", padx=(12, 4))
        self.month_filter = ttk.Entry(filters, width=16)
        self.month_filter.pack(side="left", padx=4)
        ttk.Button(filters, text="查询", command=self.refresh_all, style="Primary.TButton").pack(side="left", padx=6)
        ttk.Button(filters, text="清空", command=self.clear_filters, style="Action.TButton").pack(side="left", padx=4)
        self.status = ttk.Label(filters, text=f"数据库：{DB_PATH}", style="Muted.TLabel")
        self.status.pack(side="right", padx=10)

        panes = ttk.PanedWindow(self, orient="vertical")
        panes.pack(fill="both", expand=True, padx=14, pady=(0, 12))

        summary_frame = ttk.LabelFrame(panes, text="按店铺月份汇总", style="Surface.TLabelframe")
        panes.add(summary_frame, weight=2)
        self.summary_frozen_tree, self.summary_tree = self.make_summary_tree(summary_frame, SUMMARY_COLUMNS, ["店铺期末余额", "差异", "差异原因"], height=10)
        self.summary_frozen_tree.bind("<<TreeviewSelect>>", self.on_summary_select)
        self.summary_frozen_tree.bind("<Double-1>", lambda _e: self.edit_balance())
        self.summary_tree.bind("<<TreeviewSelect>>", self.on_summary_select)
        self.summary_tree.bind("<Double-1>", lambda _e: self.edit_balance())

        actions = ttk.Frame(summary_frame)
        actions.pack(fill="x", padx=8, pady=6)
        ttk.Button(actions, text="查看差异明细", command=self.show_difference, style="Action.TButton").pack(side="left", padx=4)
        ttk.Button(actions, text="新增调整", command=self.add_adjustment, style="Primary.TButton").pack(side="left", padx=4)
        ttk.Button(actions, text="删除选中月份流水", command=self.delete_month, style="Danger.TButton").pack(side="left", padx=4)

        detail_frame = ttk.LabelFrame(panes, text="明细流水", style="Surface.TLabelframe")
        panes.add(detail_frame, weight=3)
        self.detail_tree = self.make_tree(detail_frame, ["ID"] + RAW_COLUMNS, height=14)
        detail_actions = ttk.Frame(detail_frame)
        detail_actions.pack(fill="x", padx=8, pady=(0, 6))
        ttk.Button(detail_actions, text="编辑选中流水", command=self.edit_transaction, style="Action.TButton").pack(side="left", padx=4)
        ttk.Button(detail_actions, text="删除选中流水", command=self.delete_transaction, style="Danger.TButton").pack(side="left", padx=4)

        bottom = ttk.Frame(self, padding=(14, 0, 14, 10), style="App.TFrame")
        bottom.pack(fill="x")
        ttk.Label(bottom, text="计算口径：收入净额合计=订单实付应结+平台补贴+商家补贴+结算运费+订单退款；支出金额=已结算佣金+技术服务费；结算期末余额=期初金额+收入净额合计+支出金额+提现金额。", style="Muted.TLabel").pack(side="left")

    def make_tree(self, parent, columns, height):
        frame = ttk.Frame(parent)
        frame.pack(fill="both", expand=True, padx=8, pady=8)
        tree = ttk.Treeview(frame, columns=columns, show="headings", height=height)
        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        for col in columns:
            tree.heading(col, text=col)
            width = 130
            if col in ("商品信息",):
                width = 220
            if col in ("ID", "月份"):
                width = 70
            tree.column(col, width=width, minwidth=60, anchor="center")
        return tree

    def make_summary_tree(self, parent, base_columns, extra_columns, height):
        frozen_columns = base_columns[:2]
        scroll_columns = base_columns[2:] + extra_columns
        frame = ttk.Frame(parent)
        frame.pack(fill="both", expand=True, padx=8, pady=8)
        frozen = ttk.Treeview(frame, columns=frozen_columns, show="headings", height=height, selectmode="browse")
        scroll = ttk.Treeview(frame, columns=scroll_columns, show="headings", height=height, selectmode="browse")
        vsb = ttk.Scrollbar(frame, orient="vertical", command=lambda *args: self.sync_summary_yview(frozen, scroll, *args))
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=scroll.xview)
        frozen.configure(yscrollcommand=lambda first, last: vsb.set(first, last))
        scroll.configure(yscrollcommand=lambda first, last: vsb.set(first, last), xscrollcommand=hsb.set)
        frozen.grid(row=0, column=0, sticky="nsw")
        scroll.grid(row=0, column=1, sticky="nsew")
        vsb.grid(row=0, column=2, sticky="ns")
        hsb.grid(row=1, column=1, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(1, weight=1)
        for col in frozen_columns:
            frozen.heading(col, text=col)
            frozen.column(col, width=90, minwidth=70, anchor="center", stretch=False)
        for col in scroll_columns:
            scroll.heading(col, text=col)
            scroll.column(col, width=130, minwidth=80, anchor="center")
        frozen.bind("<MouseWheel>", lambda event: self.on_summary_mousewheel(frozen, scroll, event))
        scroll.bind("<MouseWheel>", lambda event: self.on_summary_mousewheel(frozen, scroll, event))
        return frozen, scroll

    def sync_summary_yview(self, frozen, scroll, *args):
        frozen.yview(*args)
        scroll.yview(*args)

    def on_summary_mousewheel(self, frozen, scroll, event):
        units = -1 if event.delta > 0 else 1
        frozen.yview_scroll(units, "units")
        scroll.yview_scroll(units, "units")
        return "break"

    def auto_fit_tree_columns(self, tree, max_width=260):
        for col in tree["columns"]:
            values = [str(tree.set(item, col)) for item in tree.get_children("")]
            longest = max([str(col)] + values, key=lambda value: len(value))
            width = min(max(len(longest) * 9 + 24, 70), max_width)
            if col in ("差异原因", "商品信息"):
                width = min(max(width, 180), 360)
            tree.column(col, width=width)

    def refresh_all(self):
        self.refresh_summary()
        self.refresh_details(None)

    def clear_filters(self):
        self.store_filter.delete(0, "end")
        self.month_filter.delete(0, "end")
        self.refresh_all()

    def manage_stores(self):
        StoreManageDialog(self, self.repo, self.refresh_all)

    def import_excel(self):
        store_names = [row["name"] for row in self.repo.configured_stores()]
        if not store_names:
            if messagebox.askyesno("需要配置店铺", "还没有配置店铺。是否现在添加店铺？"):
                self.manage_stores()
            return
        dialog = ImportStoreDialog(self, store_names)
        if not dialog.result:
            return
        selected_store = dialog.result
        paths = filedialog.askopenfilenames(
            title="选择资金流水 Excel",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")]
        )
        if not paths:
            return
        try:
            total_imported = total_skipped = 0
            for path in paths:
                result = self.repo.import_excel(path, selected_store=selected_store)
                total_imported += result.imported
                total_skipped += result.skipped
            self.refresh_all()
            messagebox.showinfo("导入完成", f"店铺：{selected_store}\n新增 {total_imported} 条，跳过重复/空行 {total_skipped} 条。")
        except Exception as exc:
            messagebox.showerror("导入失败", f"{exc}\n\n{traceback.format_exc(limit=2)}")

    def refresh_summary(self):
        for item in self.summary_frozen_tree.get_children():
            self.summary_frozen_tree.delete(item)
        for item in self.summary_tree.get_children():
            self.summary_tree.delete(item)
        rows = self.repo.monthly_summaries(self.store_filter.get().strip(), self.month_filter.get().strip())
        for row in rows:
            frozen_values = [row["store"], row["month_label"]]
            scroll_values = [
                money_text(row["paid_settlement"]),
                money_text(row["platform_subsidy"]), money_text(row["merchant_subsidy"]),
                money_text(row["freight"]), money_text(row["refund"]), money_text(row["income_total"]),
                money_text(row["commission"]), money_text(row["tech_fee"]),
                money_text(row["expense_amount"]), money_text(row["settlement_amount"]),
                money_text(row["withdraw_amount"]), money_text(row["opening_balance"]), money_text(row["ending_balance"]),
                "" if row["account_ending"] is None else money_text(row["account_ending"]),
                "" if row["difference"] is None else money_text(row["difference"]),
                difference_reason(row),
            ]
            tags = ("diff",) if row["difference"] is not None and row["difference"] != Decimal("0.00") else ()
            iid = f"{row['store']}|{row['month_sort']}"
            self.summary_frozen_tree.insert("", "end", values=frozen_values, tags=tags, iid=iid)
            self.summary_tree.insert("", "end", values=scroll_values, tags=tags, iid=iid)
        self.summary_frozen_tree.tag_configure("diff", background="#fff1f0")
        self.summary_tree.tag_configure("diff", background="#fff1f0")
        self.auto_fit_tree_columns(self.summary_frozen_tree, max_width=160)
        self.auto_fit_tree_columns(self.summary_tree)
        self.status.configure(text=f"数据库：{DB_PATH}    汇总 {len(rows)} 条")

    def refresh_details(self, key):
        for item in self.detail_tree.get_children():
            self.detail_tree.delete(item)
        if not key:
            return
        store, month_sort = key
        for row in self.repo.details(store, month_sort):
            values = [
                row["id"], row["store"], row["month_label"], row["transaction_time"], row["flow_id"],
                row["direction"], row["account"], money_text(row["amount"]), row["summary"],
                row["biz_type"], row["main_order"], row["sub_order"], row["after_sale"],
                row["order_time"], row["product_info"], row["product_code"], row["sale_type"],
                money_text(row["paid_settlement"]), money_text(row["platform_subsidy"]),
                money_text(row["merchant_subsidy"]), money_text(row["freight"]),
                money_text(row["refund"]), money_text(row["commission"]), money_text(row["tech_fee"]),
            ]
            self.detail_tree.insert("", "end", values=values)
        self.auto_fit_tree_columns(self.detail_tree, max_width=280)

    def current_key(self):
        selection = self.summary_tree.selection() or self.summary_frozen_tree.selection()
        if not selection:
            return None
        iid = selection[0]
        if "|" not in iid:
            return None
        return tuple(iid.split("|", 1))

    def selected_row_dict(self):
        key = self.current_key()
        if not key:
            return None
        for row in self.repo.monthly_summaries():
            if row["store"] == key[0] and row["month_sort"] == key[1]:
                return row
        return None

    def on_summary_select(self, _event=None):
        key = self.current_key()
        if key:
            iid = "|".join(key)
            if self.summary_tree.exists(iid):
                self.summary_tree.selection_set(iid)
            if self.summary_frozen_tree.exists(iid):
                self.summary_frozen_tree.selection_set(iid)
        self.refresh_details(key)

    def edit_balance(self):
        row = self.selected_row_dict()
        initial = {}
        if row:
            bal = self.repo.balance_for(row["store"], row["month_sort"])
            initial = {
                "store": row["store"],
                "month_label": row["month_label"],
                "month_sort": row["month_sort"],
                "opening_balance": row["opening_balance"],
                "account_ending": row["account_ending"] if row["account_ending"] is not None else "0",
                "note": bal["note"] if bal else "",
            }
        dialog = BalanceDialog(self, "录入/修改余额", initial)
        if dialog.result:
            self.repo.upsert_balance(
                dialog.result["store"],
                dialog.result["month_label"],
                dialog.result["month_sort"],
                dialog.result["opening_balance"],
                dialog.result["account_ending"],
                dialog.result["note"],
            )
            self.refresh_all()
            messagebox.showinfo("保存成功", "期初金额和店铺期末余额已保存。")

    def add_adjustment(self):
        row = self.selected_row_dict()
        if not row:
            messagebox.showwarning("未选择", "请先选择一个店铺月份。")
            return
        dialog = AdjustmentDialog(self, "新增差异调整", row)
        if dialog.result:
            self.repo.add_adjustment(row["store"], row["month_label"], row["month_sort"], dialog.result["target_column"], dialog.result["item"], dialog.result["amount"], dialog.result["note"])
            self.refresh_all()

    def current_transaction_id(self):
        selection = self.detail_tree.selection()
        if not selection:
            return None
        values = self.detail_tree.item(selection[0], "values")
        if not values:
            return None
        try:
            return int(values[0])
        except (TypeError, ValueError):
            return None

    def edit_transaction(self):
        tx_id = self.current_transaction_id()
        if not tx_id:
            messagebox.showwarning("未选择", "请先在明细流水中选择一条记录。")
            return
        tx = self.repo.transaction_by_id(tx_id)
        if not tx:
            messagebox.showerror("未找到", "这条流水记录不存在，可能已被删除。")
            self.refresh_all()
            return
        initial = {
            "店铺": tx["store"],
            "月份显示": tx["month_label"],
            "月份排序(YYYY-MM)": tx["month_sort"],
            "动账时间": tx["transaction_time"],
            "动账方向": tx["direction"],
            "动账账户": tx["account"],
            "动账金额": money_text(tx["amount"]).replace(",", ""),
            "动账摘要": tx["summary"],
            "业务类型": tx["biz_type"],
        }
        dialog = TransactionDialog(self, "编辑流水", initial)
        if dialog.result:
            self.repo.update_transaction(tx_id, dialog.result)
            self.refresh_all()

    def delete_transaction(self):
        tx_id = self.current_transaction_id()
        if not tx_id:
            messagebox.showwarning("未选择", "请先在明细流水中选择一条记录。")
            return
        if not messagebox.askyesno("确认删除", f"确定删除流水 ID {tx_id} 吗？"):
            return
        self.repo.delete_transaction(tx_id)
        self.refresh_all()

    def show_difference(self):
        row = self.selected_row_dict()
        if not row:
            messagebox.showwarning("未选择", "请先选择一个店铺月份。")
            return
        rows, adjs = self.repo.difference_groups(row["store"], row["month_sort"])
        win = tk.Toplevel(self)
        win.title(f"差异明细 - {row['store']} {row['month_label']}")
        win.geometry("920x560")
        text = tk.Text(win, wrap="none", font=("Consolas", 10))
        text.pack(fill="both", expand=True)
        text.insert("end", f"店铺：{row['store']}    月份：{row['month_label']} ({row['month_sort']})\n")
        text.insert("end", f"订单实付应结：{money_text(row['paid_settlement'])}\n")
        text.insert("end", f"平台补贴：{money_text(row['platform_subsidy'])}\n")
        text.insert("end", f"商家补贴：{money_text(row['merchant_subsidy'])}\n")
        text.insert("end", f"结算运费：{money_text(row['freight'])}\n")
        text.insert("end", f"订单退款：{money_text(row['refund'])}\n")
        text.insert("end", f"收入净额合计：{money_text(row['income_total'])}\n")
        text.insert("end", f"已结算佣金：{money_text(row['commission'])}\n")
        text.insert("end", f"技术服务费：{money_text(row['tech_fee'])}\n")
        text.insert("end", f"支出金额：{money_text(row['expense_amount'])}\n")
        text.insert("end", f"结算金额：{money_text(row['settlement_amount'])}\n")
        text.insert("end", f"提现金额：{money_text(row['withdraw_amount'])}\n")
        text.insert("end", f"期初金额：{money_text(row['opening_balance'])}\n")
        text.insert("end", f"结算期末余额：{money_text(row['ending_balance'])}\n")
        text.insert("end", f"店铺期末余额：{'' if row['account_ending'] is None else money_text(row['account_ending'])}\n")
        text.insert("end", f"差异：{'' if row['difference'] is None else money_text(row['difference'])}\n\n")
        text.insert("end", "差异原因说明：\n")
        text.insert("end", f"{difference_reason(row)}\n")
        text.insert("end", "核对公式：差异 = 结算期末余额 - 店铺期末余额；结算期末余额 = 期初金额 + 收入净额合计 + 支出金额 + 提现金额。\n\n")
        text.insert("end", "按动账方向/摘要汇总：\n")
        text.insert("end", "方向\t摘要\t笔数\t动账金额\t订单实付应结\t平台补贴\t商家补贴\t结算运费\t订单退款\t收入净额合计\t已结算佣金\t技术服务费\t支出金额\n")
        for r in rows:
            text.insert("end", f"{r['direction']}\t{r['summary']}\t{r['count']}\t{money_text(r['amount'])}\t{money_text(r['paid_settlement'])}\t{money_text(r['platform_subsidy'])}\t{money_text(r['merchant_subsidy'])}\t{money_text(r['freight'])}\t{money_text(r['refund'])}\t{money_text(r['income_total'])}\t{money_text(r['commission'])}\t{money_text(r['tech_fee'])}\t{money_text(r['expense_amount'])}\n")
        text.insert("end", "\n手工调整记录（选择具体列会参与汇总；选择备查不影响汇总）：\n")
        if adjs:
            for a in adjs:
                text.insert("end", f"#{a['id']} [{a['target_column'] or '备查（不影响汇总）'}] {a['item']}\t{money_text(a['amount'])}\t{a['note'] or ''}\n")
        else:
            text.insert("end", "无\n")
        text.configure(state="disabled")

    def delete_month(self):
        row = self.selected_row_dict()
        if not row:
            messagebox.showwarning("未选择", "请先选择要删除的店铺月份。")
            return
        if not messagebox.askyesno("确认删除", f"确定删除 {row['store']} {row['month_label']} 的全部流水和手工调整吗？余额记录默认保留。"):
            return
        self.repo.delete_month(row["store"], row["month_sort"], delete_balance=False)
        self.refresh_all()

    def export_summary(self):
        path = filedialog.asksaveasfilename(
            title="保存汇总CSV",
            defaultextension=".csv",
            filetypes=[("CSV 文件", "*.csv")]
        )
        if not path:
            return
        rows = self.repo.monthly_summaries(self.store_filter.get().strip(), self.month_filter.get().strip())
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(SUMMARY_COLUMNS + ["店铺期末余额", "差异", "差异原因"])
            for row in rows:
                writer.writerow([
                    row["store"], row["month_label"], row["paid_settlement"],
                    row["platform_subsidy"], row["merchant_subsidy"], row["freight"], row["refund"],
                    row["income_total"], row["commission"], row["tech_fee"], row["expense_amount"],
                    row["settlement_amount"], row["withdraw_amount"], row["opening_balance"], row["ending_balance"],
                    "" if row["account_ending"] is None else row["account_ending"],
                    "" if row["difference"] is None else row["difference"],
                    difference_reason(row),
                ])
        messagebox.showinfo("导出完成", f"已保存：{path}")


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
