# -*- coding: utf-8 -*-
"""Qt 主界面模块。

本文件负责 Windows 桌面端交互：工具栏、弹窗、汇总表、明细表、
筛选分页、导出和全局设置。数据库读写、导入解析、金额计算统一放在
payment_recon_qt.Repository 中，避免界面代码直接拼业务 SQL。
"""
import json
import os
import sys
import traceback
from datetime import datetime
from decimal import Decimal
from pathlib import Path

LOCAL_PYSIDE = Path(__file__).resolve().parents[2] / "work" / "pyside6_pkg"
if LOCAL_PYSIDE.exists():
    sys.path.insert(0, str(LOCAL_PYSIDE))

from PySide6.QtCore import Qt
from PySide6.QtGui import QAction, QColor, QFontMetrics, QIcon, QKeySequence
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMenu,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QSplitter,
    QSpinBox,
    QSystemTrayIcon,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QToolBar,
    QVBoxLayout,
    QWidget,
)
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter

from payment_recon_qt import (
    ADJUSTABLE_COLUMNS,
    DB_PATH,
    RAW_COLUMNS,
    REPORT_TYPES,
    SUMMARY_COLUMNS,
    SUMMARY_EXTRA_COLUMNS,
    Repository,
    clean_cell_text,
    difference_reason,
    month_range,
    money,
    money_text,
    parse_month_filter,
)


APP_TITLE = "财务第三方支付核对 - Qt版"
APP_VERSION = "1.0.4"


# 界面统一显示“年月”，但底层旧配置仍可能叫“月份”。
# 这里集中维护显示名和数据名的互相转换。
DISPLAY_HEADER_MAP = {
    "月份": "年月",
    "月份排序": "年月排序",
    "月份排序(YYYY-MM)": "年月排序(YYYY-MM)",
    "月份显示": "年月显示",
    "月份区间": "年月区间",
}

DATA_HEADER_MAP = {display: data for data, display in DISPLAY_HEADER_MAP.items()}


def display_header(name):
    """把底层字段名转换为界面展示名。"""
    return DISPLAY_HEADER_MAP.get(name, name)


def data_header(name):
    """把界面展示名转换为底层字段名。"""
    return DATA_HEADER_MAP.get(name, name)


def localize_buttons(buttons, ok_text="确定", cancel_text="退出", close_text="关闭"):
    """统一弹窗右下角按钮文案，避免默认 OK/Cancel 出现在中文界面。"""
    ok_button = buttons.button(QDialogButtonBox.Ok)
    if ok_button:
        ok_button.setText(ok_text)
    cancel_button = buttons.button(QDialogButtonBox.Cancel)
    if cancel_button:
        cancel_button.setText(cancel_text)
    close_button = buttons.button(QDialogButtonBox.Close)
    if close_button:
        close_button.setText(close_text)


def resource_path(name):
    """兼容源码运行和 PyInstaller onefile 运行时的资源路径。"""
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return base / name


def table_item(value, align=Qt.AlignCenter):
    """创建只读表格单元格，统一对齐和编辑状态。"""
    item = QTableWidgetItem(str(value))
    item.setTextAlignment(align)
    item.setFlags(item.flags() ^ Qt.ItemIsEditable)
    return item


def editable_table_item(value, align=Qt.AlignLeft | Qt.AlignVCenter):
    """创建可直接编辑的表格单元格，用在参数配置页。"""
    item = QTableWidgetItem(str(value))
    item.setTextAlignment(align)
    item.setFlags(item.flags() | Qt.ItemIsEditable)
    return item


class CopyableTableWidget(QTableWidget):
    """支持 Ctrl+C 复制选中单元格的参数配置表格。"""

    def keyPressEvent(self, event):
        if event.matches(QKeySequence.Copy):
            self.copy_selection()
            return
        super().keyPressEvent(event)

    def copy_selection(self):
        ranges = self.selectedRanges()
        if not ranges:
            item = self.currentItem()
            if item:
                QApplication.clipboard().setText(item.text())
            return
        lines = []
        for selected in ranges:
            for row in range(selected.topRow(), selected.bottomRow() + 1):
                values = []
                for col in range(selected.leftColumn(), selected.rightColumn() + 1):
                    item = self.item(row, col)
                    values.append(item.text() if item else "")
                lines.append("\t".join(values))
        QApplication.clipboard().setText("\n".join(lines))


class FieldTableWidget(CopyableTableWidget):
    """字段列表表格：Delete 删除确认，拖动只移动行，不覆盖单元格。"""

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key_Delete, Qt.Key_Backspace):
            row = self.currentRow()
            item = self.item(row, 0) if row >= 0 else None
            if item and QMessageBox.question(self, "确认删除字段", f"确定删除字段：{item.text()}？") == QMessageBox.Yes:
                self.removeRow(row)
            return
        super().keyPressEvent(event)

    def dropEvent(self, event):
        source_row = self.currentRow()
        if source_row < 0 or source_row >= self.rowCount():
            event.ignore()
            return
        pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
        target_row = self.rowAt(pos.y())
        if target_row < 0:
            target_row = self.rowCount()
        else:
            rect = self.visualRect(self.model().index(target_row, 0))
            if pos.y() > rect.center().y():
                target_row += 1
        if target_row == source_row or target_row == source_row + 1:
            event.accept()
            return
        text = self.item(source_row, 0).text() if self.item(source_row, 0) else ""
        self.removeRow(source_row)
        if target_row > source_row:
            target_row -= 1
        target_row = max(0, min(target_row, self.rowCount()))
        self.insertRow(target_row)
        self.setItem(target_row, 0, editable_table_item(text))
        self.selectRow(target_row)
        event.accept()


class BalanceDialog(QDialog):
    """录入或修改期初金额、店铺期末余额的弹窗。"""

    def __init__(self, parent, initial):
        super().__init__(parent)
        self.setWindowTitle("录入/修改余额")
        self.resize(460, 280)
        self.result_data = None

        self.store = QLineEdit(str(initial.get("store", "")))
        self.report_type = QComboBox()
        self.report_type.addItems(REPORT_TYPES)
        self.report_type.setCurrentText(str(initial.get("report_type", "已结算")))
        self.month_label = QLineEdit(str(initial.get("month_label", "")))
        self.month_sort = QLineEdit(str(initial.get("month_sort", "")))
        self.opening = QLineEdit(str(initial.get("opening_balance", "0")))
        self.account_ending = QLineEdit(str(initial.get("account_ending", "0")))
        self.note = QLineEdit(str(initial.get("note", "")))

        form = QFormLayout()
        form.addRow("店铺", self.store)
        form.addRow("报表类型", self.report_type)
        form.addRow("年月显示", self.month_label)
        form.addRow("年月排序(YYYY-MM)", self.month_sort)
        form.addRow("期初金额", self.opening)
        form.addRow("店铺期末余额", self.account_ending)
        form.addRow("备注", self.note)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        localize_buttons(buttons, "保存余额", "退出")
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)

        layout = QVBoxLayout(self)
        layout.addLayout(form)
        layout.addWidget(buttons)

    def accept(self):
        if not self.store.text().strip() or not self.month_sort.text().strip():
            QMessageBox.warning(self, "请补充信息", "店铺和年月排序不能为空。")
            return
        self.result_data = {
            "store": self.store.text().strip(),
            "report_type": self.report_type.currentText(),
            "month_label": self.month_label.text().strip() or self.month_sort.text().strip(),
            "month_sort": self.month_sort.text().strip(),
            "opening_balance": self.opening.text().strip() or "0",
            "account_ending": self.account_ending.text().strip() or "0",
            "note": self.note.text().strip(),
        }
        super().accept()


class AdjustmentDialog(QDialog):
    """新增手工调整记录，并说明调整影响哪一个汇总字段。"""

    def __init__(self, parent, row):
        super().__init__(parent)
        self.setWindowTitle("新增差异调整")
        self.resize(480, 300)
        self.result_data = None

        self.target = QComboBox()
        self.target.addItems(ADJUSTABLE_COLUMNS)
        self.item = QLineEdit("差异调整")
        self.amount = QLineEdit("0")
        self.note = QLineEdit("")
        info = QLabel(f"{row['store']} / {row['month_label']}    当前差异：{'' if row['difference'] is None else money_text(row['difference'])}")
        info.setObjectName("mutedLabel")

        form = QFormLayout()
        form.addRow("调整列", self.target)
        form.addRow("调整事项", self.item)
        form.addRow("调整金额", self.amount)
        form.addRow("说明", self.note)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        localize_buttons(buttons, "保存调整", "退出")
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)

        layout = QVBoxLayout(self)
        layout.addWidget(info)
        layout.addLayout(form)
        layout.addWidget(buttons)

    def accept(self):
        if not self.item.text().strip():
            QMessageBox.warning(self, "请补充信息", "调整事项不能为空。")
            return
        self.result_data = {
            "target_column": self.target.currentText(),
            "item": self.item.text().strip(),
            "amount": self.amount.text().strip() or "0",
            "note": self.note.text().strip(),
        }
        super().accept()


class TransactionDialog(QDialog):
    """编辑单条明细流水的常用字段。"""

    FIELDS = ["店铺", "月份显示", "月份排序(YYYY-MM)", "动账时间", "动账方向", "动账账户", "动账金额", "动账摘要", "业务类型"]

    def __init__(self, parent, initial):
        super().__init__(parent)
        self.setWindowTitle("编辑流水")
        self.resize(520, 420)
        self.result_data = None
        self.edits = {}

        form = QFormLayout()
        for field in self.FIELDS:
            edit = QLineEdit(str(initial.get(field, "")))
            self.edits[field] = edit
            form.addRow(display_header(field), edit)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        localize_buttons(buttons, "保存流水", "退出")
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)

        layout = QVBoxLayout(self)
        layout.addLayout(form)
        layout.addWidget(buttons)

    def accept(self):
        if not self.edits["店铺"].text().strip() or not self.edits["月份排序(YYYY-MM)"].text().strip():
            QMessageBox.warning(self, "请补充信息", "店铺和年月排序不能为空。")
            return
        self.result_data = {field: edit.text().strip() for field, edit in self.edits.items()}
        super().accept()


class MonthSelectDialog(QDialog):
    """选择首页查询年月区间的弹窗。"""

    def __init__(self, parent, months, selected):
        super().__init__(parent)
        self.setWindowTitle("选择年月区间")
        self.resize(520, 220)
        self.result_months = None
        years = [str(year) for year in range(2025, 2100)]
        month_numbers = [f"{i:02d}" for i in range(1, 13)]
        self.start_year_combo = QComboBox()
        self.start_month_combo = QComboBox()
        self.end_year_combo = QComboBox()
        self.end_month_combo = QComboBox()
        self.start_year_combo.addItems(years)
        self.end_year_combo.addItems(years)
        self.start_month_combo.addItems(month_numbers)
        self.end_month_combo.addItems(month_numbers)
        selected = selected or []
        current_month = datetime.now().strftime("%Y-%m")
        start_value = selected[0] if selected else current_month
        end_value = selected[-1] if selected else current_month
        self.set_combo_value(self.start_year_combo, start_value[:4])
        self.set_combo_value(self.start_month_combo, start_value[5:7])
        self.set_combo_value(self.end_year_combo, end_value[:4])
        self.set_combo_value(self.end_month_combo, end_value[5:7])

        form = QFormLayout()
        start_row = QHBoxLayout()
        start_row.addWidget(self.start_year_combo)
        start_row.addWidget(QLabel("年"))
        start_row.addWidget(self.start_month_combo)
        start_row.addWidget(QLabel("月"))
        end_row = QHBoxLayout()
        end_row.addWidget(self.end_year_combo)
        end_row.addWidget(QLabel("年"))
        end_row.addWidget(self.end_month_combo)
        end_row.addWidget(QLabel("月"))
        form.addRow("开始年月", start_row)
        form.addRow("结束年月", end_row)

        self.preview = QLabel("")
        self.preview.setObjectName("mutedLabel")
        for combo in (self.start_year_combo, self.start_month_combo, self.end_year_combo, self.end_month_combo):
            combo.currentIndexChanged.connect(self.update_preview)
        self.update_preview()

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        localize_buttons(buttons, "选择年月", "退出")
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("选择开始和结束年月后，会自动包含中间所有年月。"))
        layout.addLayout(form)
        layout.addWidget(self.preview)
        layout.addWidget(buttons)

    def set_combo_value(self, combo, value):
        index = combo.findText(value)
        if index >= 0:
            combo.setCurrentIndex(index)

    def combo_month(self, year_combo, month_combo):
        return f"{year_combo.currentText()}-{month_combo.currentText()}"

    def selected_range(self):
        start = self.combo_month(self.start_year_combo, self.start_month_combo)
        end = self.combo_month(self.end_year_combo, self.end_month_combo)
        if start > end:
            start, end = end, start
        return month_range(start, end)

    def update_preview(self):
        months = self.selected_range()
        self.preview.setText("已选择：" + (", ".join(months) if months else "无可选年月"))

    def accept(self):
        self.result_months = self.selected_range()
        super().accept()


class ImportOptionsDialog(QDialog):
    """导入流水前的一站式设置弹窗：店铺、报表类型、归属年月。"""

    def __init__(self, parent, stores, current_store="", initial_month=None):
        super().__init__(parent)
        self.setWindowTitle("导入流水设置")
        self.resize(440, 220)
        current = initial_month or datetime.now().strftime("%Y-%m")
        self.result_data = None
        self.store_combo = QComboBox()
        self.store_combo.addItems(stores)
        if current_store in stores:
            self.store_combo.setCurrentText(current_store)
        self.report_type_combo = QComboBox()
        self.report_type_combo.addItems(REPORT_TYPES)
        self.year_combo = QComboBox()
        self.month_combo = QComboBox()
        self.year_combo.addItems([str(year) for year in range(2025, 2100)])
        self.month_combo.addItems([f"{month:02d}" for month in range(1, 13)])
        year_index = self.year_combo.findText(current[:4])
        month_index = self.month_combo.findText(current[5:7])
        if year_index >= 0:
            self.year_combo.setCurrentIndex(year_index)
        if month_index >= 0:
            self.month_combo.setCurrentIndex(month_index)

        month_row = QHBoxLayout()
        month_row.addWidget(self.year_combo)
        month_row.addWidget(QLabel("年"))
        month_row.addWidget(self.month_combo)
        month_row.addWidget(QLabel("月"))

        form = QFormLayout()
        form.addRow("店铺", self.store_combo)
        form.addRow("报表类型", self.report_type_combo)
        form.addRow("归属年月", month_row)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        localize_buttons(buttons, "开始导入", "退出")
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("请选择本次导入流水的店铺、报表类型和归属年月。"))
        layout.addLayout(form)
        layout.addWidget(buttons)

    def accept(self):
        self.result_data = {
            "store": self.store_combo.currentText().strip(),
            "report_type": self.report_type_combo.currentText().strip(),
            "import_month": f"{self.year_combo.currentText()}-{self.month_combo.currentText()}",
        }
        super().accept()


class GlobalSettingsDialog(QDialog):
    """全局设置弹窗，例如开机自启动和关闭到托盘。"""

    def __init__(self, parent, settings):
        super().__init__(parent)
        self.setWindowTitle("全局设置")
        self.resize(420, 180)
        self.result_data = None
        self.startup = QCheckBox("开机自动启动")
        self.close_to_tray = QCheckBox("点击 X 时最小化到托盘，不直接退出")
        self.startup.setChecked(bool(settings.get("startup_enabled")))
        self.close_to_tray.setChecked(bool(settings.get("close_to_tray")))

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        localize_buttons(buttons, "保存设置", "退出")
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)

        layout = QVBoxLayout(self)
        layout.addWidget(self.startup)
        layout.addWidget(self.close_to_tray)
        layout.addWidget(buttons)

    def accept(self):
        self.result_data = {
            "startup_enabled": self.startup.isChecked(),
            "close_to_tray": self.close_to_tray.isChecked(),
        }
        super().accept()


class BackupRestoreDialog(QDialog):
    """备份和还原店铺数据库、主配置、全局设置的弹窗。"""

    def __init__(self, parent, repo):
        super().__init__(parent)
        self.repo = repo
        self.restore_manifest = None
        self.setWindowTitle("备份还原")
        self.resize(760, 520)

        tabs = QTabWidget()
        tabs.addTab(self.build_backup_tab(), "备份")
        tabs.addTab(self.build_restore_tab(), "还原")

        close_btn = QDialogButtonBox(QDialogButtonBox.Close)
        localize_buttons(close_btn, close_text="关闭")
        close_btn.rejected.connect(self.reject)
        layout = QVBoxLayout(self)
        layout.addWidget(tabs)
        self.progress_label = QLabel("等待操作")
        self.progress_label.setObjectName("statusPill")
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        layout.addWidget(self.progress_label)
        layout.addWidget(self.progress_bar)
        layout.addWidget(close_btn)

    def set_progress(self, message, value=0, total=100):
        total = max(1, int(total or 1))
        value = max(0, min(int(value or 0), total))
        self.progress_label.setText(message)
        self.progress_bar.setRange(0, total)
        self.progress_bar.setValue(value)
        QApplication.processEvents()

    def build_store_table(self, rows):
        table = QTableWidget(0, 3)
        table.setHorizontalHeaderLabels(["选择", "店铺", "状态"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.verticalHeader().setVisible(False)
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            check = QTableWidgetItem("")
            check.setFlags(check.flags() | Qt.ItemIsUserCheckable)
            check.setCheckState(Qt.Checked)
            table.setItem(r, 0, check)
            table.setItem(r, 1, QTableWidgetItem(row.get("name", "")))
            table.setItem(r, 2, QTableWidgetItem("启用" if int(row.get("active", 1) or 0) else "停用"))
        table.setColumnWidth(0, 70)
        return table

    def checked_store_names(self, table):
        names = []
        for row in range(table.rowCount()):
            item = table.item(row, 0)
            name_item = table.item(row, 1)
            if item and name_item and item.checkState() == Qt.Checked:
                names.append(name_item.text())
        return names

    def build_backup_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        info = QLabel("选择要备份的店铺，并勾选备份内容。可以只备份数据库、只备份配置，或两者都备份。")
        info.setWordWrap(True)
        layout.addWidget(info)
        rows = [dict(row) for row in self.repo.configured_stores(include_inactive=True)]
        self.backup_store_table = self.build_store_table(rows)
        layout.addWidget(self.backup_store_table, 1)

        self.backup_database_check = QCheckBox("备份店铺数据库")
        self.backup_config_check = QCheckBox("备份全部配置（店铺配置和全局设置）")
        self.backup_database_check.setChecked(True)
        self.backup_config_check.setChecked(True)
        layout.addWidget(self.backup_database_check)
        layout.addWidget(self.backup_config_check)
        self.backup_settings_check = self.backup_config_check
        self.backup_settings_check.setChecked(True)

        path_row = QHBoxLayout()
        self.backup_path_edit = QLineEdit()
        default_name = f"财务核对备份_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        self.backup_path_edit.setText(str(Path.home() / "Desktop" / default_name))
        browse = QPushButton("选择位置")
        browse.clicked.connect(self.choose_backup_path)
        path_row.addWidget(QLabel("保存到"))
        path_row.addWidget(self.backup_path_edit, 1)
        path_row.addWidget(browse)
        layout.addLayout(path_row)

        self.backup_btn = QPushButton("开始备份")
        self.backup_btn.clicked.connect(self.run_backup)
        layout.addWidget(self.backup_btn)
        return widget

    def build_restore_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        info = QLabel("选择备份文件后，可勾选要还原的店铺和内容。还原数据库会覆盖同名店铺本地数据库。")
        info.setWordWrap(True)
        layout.addWidget(info)

        path_row = QHBoxLayout()
        self.restore_path_edit = QLineEdit()
        browse = QPushButton("选择备份文件")
        browse.clicked.connect(self.choose_restore_path)
        path_row.addWidget(QLabel("备份文件"))
        path_row.addWidget(self.restore_path_edit, 1)
        path_row.addWidget(browse)
        layout.addLayout(path_row)

        self.restore_table_box = QWidget()
        self.restore_table_layout = QVBoxLayout(self.restore_table_box)
        self.restore_table_layout.setContentsMargins(0, 0, 0, 0)
        self.restore_store_table = self.build_store_table([])
        self.restore_table_layout.addWidget(self.restore_store_table)
        layout.addWidget(self.restore_table_box, 1)
        self.restore_database_check = QCheckBox("还原店铺数据库")
        self.restore_config_check = QCheckBox("还原全部配置（店铺配置和全局设置）")
        self.restore_database_check.setChecked(True)
        self.restore_config_check.setChecked(True)
        layout.addWidget(self.restore_database_check)
        layout.addWidget(self.restore_config_check)
        self.restore_settings_check = self.restore_config_check

        self.restore_btn = QPushButton("开始还原")
        self.restore_btn.clicked.connect(self.run_restore)
        layout.addWidget(self.restore_btn)
        return widget

    def choose_backup_path(self):
        path, _ = QFileDialog.getSaveFileName(self, "保存备份文件", self.backup_path_edit.text(), "备份文件 (*.zip)")
        if path:
            if not path.lower().endswith(".zip"):
                path += ".zip"
            self.backup_path_edit.setText(path)

    def choose_restore_path(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择备份文件", "", "备份文件 (*.zip)")
        if not path:
            return
        self.restore_path_edit.setText(path)
        try:
            self.restore_manifest = self.repo.inspect_backup(path)
            old_table = self.restore_store_table
            self.restore_table_layout.removeWidget(old_table)
            old_table.deleteLater()
            rows = self.restore_manifest.get("stores", [])
            self.restore_store_table = self.build_store_table(rows)
            self.restore_table_layout.addWidget(self.restore_store_table)
        except Exception as exc:
            QMessageBox.warning(self, "读取失败", f"无法读取备份文件：{exc}")

    def run_backup(self):
        stores = self.checked_store_names(self.backup_store_table)
        if not stores:
            QMessageBox.warning(self, "未选择店铺", "请至少选择一个店铺。")
            return
        path = self.backup_path_edit.text().strip()
        if not path:
            QMessageBox.warning(self, "未选择路径", "请选择备份保存路径。")
            return
        include_database = self.backup_database_check.isChecked()
        include_config = self.backup_config_check.isChecked()
        if not include_database and not include_config:
            QMessageBox.warning(self, "未选择内容", "请至少选择备份店铺数据库或备份全部配置。")
            return
        try:
            self.backup_btn.setEnabled(False)
            self.set_progress("正在准备备份...", 0, max(1, len(stores) + 2))
            result = self.repo.backup_data(
                path,
                stores,
                include_settings=include_config,
                include_databases=include_database,
                include_configs=include_config,
                progress_callback=self.set_progress,
            )
            content = "、".join([text for enabled, text in ((include_database, "店铺数据库"), (include_config, "全部配置")) if enabled])
            self.set_progress(f"备份完成：{result['stores']} 个店铺", 100, 100)
            QMessageBox.information(self, "备份完成", f"已备份 {result['stores']} 个店铺。\n内容：{content}\n文件：{result['path']}")
        except Exception as exc:
            self.set_progress(f"备份失败：{exc}", 0, 100)
            QMessageBox.critical(self, "备份失败", str(exc))
        finally:
            self.backup_btn.setEnabled(True)

    def run_restore(self):
        path = self.restore_path_edit.text().strip()
        if not path:
            QMessageBox.warning(self, "未选择备份", "请先选择备份文件。")
            return
        stores = self.checked_store_names(self.restore_store_table)
        if not stores:
            QMessageBox.warning(self, "未选择店铺", "请至少选择一个要还原的店铺。")
            return
        restore_database = self.restore_database_check.isChecked()
        restore_config = self.restore_config_check.isChecked()
        if not restore_database and not restore_config:
            QMessageBox.warning(self, "未选择内容", "请至少选择还原店铺数据库或还原全部配置。")
            return
        content = "、".join([text for enabled, text in ((restore_database, "店铺数据库"), (restore_config, "全部配置")) if enabled])
        if QMessageBox.question(self, "确认还原", f"将还原：{content}。\n还原数据库会覆盖同名店铺本地数据库，是否继续？") != QMessageBox.Yes:
            return
        try:
            self.restore_btn.setEnabled(False)
            self.set_progress("正在准备还原...", 0, max(1, len(stores) + 2))
            result = self.repo.restore_data(
                path,
                stores,
                restore_settings=restore_config,
                restore_databases=restore_database,
                restore_configs=restore_config,
                progress_callback=self.set_progress,
            )
            self.set_progress(f"还原完成：{result['stores']} 个店铺", 100, 100)
            QMessageBox.information(self, "还原完成", f"已还原 {result['stores']} 个店铺。\n内容：{content}")
            self.accept()
        except Exception as exc:
            self.set_progress(f"还原失败：{exc}", 0, 100)
            QMessageBox.critical(self, "还原失败", str(exc))
        finally:
            self.restore_btn.setEnabled(True)


class StoreDialog(QDialog):
    """店铺配置弹窗，支持直接编辑店铺名称、备注和状态。"""

    def __init__(self, parent, repo):
        super().__init__(parent)
        self.repo = repo
        self.setWindowTitle("店铺配置")
        self.resize(560, 420)
        self.original_names = []

        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(["店铺", "备注", "状态"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)

        add_btn = QPushButton("新增/启用店铺")
        save_btn = QPushButton("保存表格修改")
        disable_btn = QPushButton("停用选中店铺")
        delete_btn = QPushButton("删除选中店铺")
        add_btn.clicked.connect(self.add_store)
        save_btn.clicked.connect(self.save_edits)
        disable_btn.clicked.connect(self.disable_store)
        delete_btn.clicked.connect(self.delete_store)

        bar = QHBoxLayout()
        bar.addWidget(add_btn)
        bar.addWidget(save_btn)
        bar.addWidget(disable_btn)
        bar.addWidget(delete_btn)
        bar.addStretch()

        layout = QVBoxLayout(self)
        layout.addWidget(self.table)
        layout.addLayout(bar)
        self.refresh()

    def refresh(self):
        rows = self.repo.configured_stores(include_inactive=True)
        self.original_names = [row["name"] for row in rows]
        self.table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            self.table.setItem(r, 0, QTableWidgetItem(row["name"]))
            self.table.setItem(r, 1, QTableWidgetItem(row["note"] or ""))
            self.table.setItem(r, 2, QTableWidgetItem("启用" if row["active"] else "停用"))
        self.table.resizeColumnsToContents()

    def add_store(self):
        name, ok = QInputDialog.getText(self, "新增店铺", "店铺名称")
        if not ok or not name.strip():
            return
        note, _ = QInputDialog.getText(self, "店铺备注", "备注（可为空）")
        try:
            self.repo.add_store(name.strip(), note.strip())
            self.refresh()
        except Exception as exc:
            QMessageBox.critical(self, "保存失败", str(exc))

    def disable_store(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "未选择", "请先选择一个店铺。")
            return
        name = self.table.item(row, 0).text()
        if QMessageBox.question(self, "确认停用", f"确定停用店铺：{name}？") != QMessageBox.Yes:
            return
        self.repo.delete_store(name)
        self.refresh()

    def delete_store(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "未选择", "请先选择一个店铺。")
            return
        name = self.table.item(row, 0).text().strip() if self.table.item(row, 0) else ""
        if not name:
            QMessageBox.warning(self, "未选择", "店铺名称不能为空。")
            return
        message = (
            f"确定永久删除店铺：{name}？\n\n"
            "删除后会移除该店铺配置、历史流水、余额、调整记录和独立数据库文件。"
        )
        if QMessageBox.question(self, "确认删除店铺", message) != QMessageBox.Yes:
            return
        try:
            result = self.repo.delete_store_permanently(name, delete_database=True)
            self.refresh()
            if result.get("database_deleted", True):
                QMessageBox.information(self, "删除完成", f"已删除店铺：{name}")
            else:
                QMessageBox.warning(
                    self,
                    "删除完成",
                    f"已删除店铺：{name}\n数据库文件当前被占用，关闭程序后可清理：{result.get('db_file', '')}",
                )
        except Exception as exc:
            QMessageBox.critical(self, "删除失败", str(exc))

    def save_edits(self):
        try:
            for row in range(self.table.rowCount()):
                old_name = self.original_names[row] if row < len(self.original_names) else ""
                name = self.table.item(row, 0).text().strip() if self.table.item(row, 0) else ""
                note = self.table.item(row, 1).text().strip() if self.table.item(row, 1) else ""
                status = self.table.item(row, 2).text().strip() if self.table.item(row, 2) else "启用"
                active = 0 if status in ("停用", "0", "否", "禁用") else 1
                self.repo.update_store(old_name or name, name, note, active)
            self.refresh()
            QMessageBox.information(self, "保存成功", "店铺资料已保存。")
        except Exception as exc:
            QMessageBox.critical(self, "保存失败", str(exc))


class FieldListEditor(QWidget):
    """参数配置中的字段列表编辑器，支持增删、排序和冻结/活动区互移。"""

    def __init__(self, values, title):
        super().__init__()
        self.table = FieldTableWidget(0, 1)
        self.table.setHorizontalHeaderLabels([title])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.SelectedClicked | QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        self.table.setDragDropMode(QTableWidget.InternalMove)
        self.table.setDragDropOverwriteMode(False)
        self.table.setDefaultDropAction(Qt.MoveAction)
        self.table.setUpdatesEnabled(False)
        for index, value in enumerate(values or [], 1):
            self.add_value(value)
            if index % max(100, int(batch_size or 1000)) == 0:
                QApplication.processEvents()
        self.table.setUpdatesEnabled(True)

        add_btn = QPushButton("新增字段")
        del_btn = QPushButton("删除字段")
        up_btn = QPushButton("上移")
        down_btn = QPushButton("下移")
        top_btn = QPushButton("置顶")
        add_btn.clicked.connect(self.add_dialog)
        del_btn.clicked.connect(self.delete_selected)
        up_btn.clicked.connect(lambda: self.move_selected(-1))
        down_btn.clicked.connect(lambda: self.move_selected(1))
        top_btn.clicked.connect(self.move_top)

        bar = QHBoxLayout()
        for btn in (add_btn, del_btn, top_btn, up_btn, down_btn):
            bar.addWidget(btn)
        bar.addStretch()
        layout = QVBoxLayout(self)
        layout.addWidget(self.table, 1)
        layout.addLayout(bar)

    def add_value(self, value):
        value = str(value or "").strip()
        if not value:
            return
        row = self.table.rowCount()
        self.table.insertRow(row)
        self.table.setItem(row, 0, editable_table_item(value))

    def add_dialog(self):
        value, ok = QInputDialog.getText(self, "新增字段", "字段名称")
        if ok:
            self.add_value(value)

    def edit_dialog(self):
        row = self.table.currentRow()
        if row < 0:
            return
        old = self.table.item(row, 0).text()
        value, ok = QInputDialog.getText(self, "修改字段", "字段名称", text=old)
        if ok and value.strip():
            self.table.setItem(row, 0, editable_table_item(value.strip()))

    def delete_selected(self):
        row = self.table.currentRow()
        item = self.table.item(row, 0) if row >= 0 else None
        if item and QMessageBox.question(self, "确认删除字段", f"确定删除字段：{item.text()}？") == QMessageBox.Yes:
            self.table.removeRow(row)

    def move_selected(self, step):
        row = self.table.currentRow()
        target = row + step
        if row < 0 or target < 0 or target >= self.table.rowCount():
            return
        current = self.table.takeItem(row, 0).text()
        other = self.table.takeItem(target, 0).text()
        self.table.setItem(row, 0, editable_table_item(other))
        self.table.setItem(target, 0, editable_table_item(current))
        self.table.selectRow(target)

    def move_top(self):
        row = self.table.currentRow()
        if row <= 0:
            return
        current = self.table.takeItem(row, 0).text()
        self.table.removeRow(row)
        self.table.insertRow(0)
        self.table.setItem(0, 0, editable_table_item(current))
        self.table.selectRow(0)

    def values(self):
        result = []
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if item and item.text().strip() and item.text().strip() not in result:
                result.append(item.text().strip())
        return result

    def selected_value(self):
        row = self.table.currentRow()
        if row < 0 or not self.table.item(row, 0):
            return ""
        return self.table.item(row, 0).text().strip()

    def remove_value(self, value):
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if item and item.text().strip() == value:
                self.table.removeRow(row)
                return


class FormulaEditor(QWidget):
    """参数配置中的计算口径编辑器。"""

    def __init__(self, formula_note, batch_size=1000):
        super().__init__()
        self.table = CopyableTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(["汇总字段", "计算公式"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.SelectedClicked | QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        self.table.setUpdatesEnabled(False)
        for index, line in enumerate(str(formula_note or "").splitlines(), 1):
            if "=" in line:
                target, expr = line.split("=", 1)
                self.add_row(target.strip(), expr.strip())
            if index % max(100, int(batch_size or 1000)) == 0:
                QApplication.processEvents()
        self.table.setUpdatesEnabled(True)

        add_btn = QPushButton("新增口径")
        del_btn = QPushButton("删除口径")
        add_btn.clicked.connect(lambda: self.add_row("", ""))
        del_btn.clicked.connect(self.delete_selected)
        bar = QHBoxLayout()
        bar.addWidget(add_btn)
        bar.addWidget(del_btn)
        bar.addStretch()
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("公式里可引用汇总字段名称，例如：收入净额合计=订单实付应结+平台补贴+商家补贴+结算运费+订单退款"))
        layout.addWidget(self.table, 1)
        layout.addLayout(bar)

    def add_row(self, target, expr):
        row = self.table.rowCount()
        self.table.insertRow(row)
        self.table.setItem(row, 0, QTableWidgetItem(target))
        self.table.setItem(row, 1, QTableWidgetItem(expr))

    def delete_selected(self):
        row = self.table.currentRow()
        if row >= 0 and QMessageBox.question(self, "确认删除口径", "确定删除选中的计算口径？") == QMessageBox.Yes:
            self.table.removeRow(row)

    def text(self):
        lines = []
        for row in range(self.table.rowCount()):
            target = self.table.item(row, 0).text().strip() if self.table.item(row, 0) else ""
            expr = self.table.item(row, 1).text().strip() if self.table.item(row, 1) else ""
            if target and expr:
                lines.append(f"{target}={expr}")
        return "\n".join(lines)


class StoreConfigDialog(QDialog):
    """单店铺参数配置弹窗：原始字段、汇总字段、计算口径和分页行数。"""

    def __init__(self, parent, repo, store):
        super().__init__(parent)
        self.repo = repo
        self.store = store
        self.initial_state = None
        self.saved = False
        self.setWindowTitle(f"参数配置 - {store}")
        self.resize(920, 680)
        config = repo.store_config(store)
        summary_columns = list(config["summary_columns"])
        frozen_columns = [col for col in config.get("frozen_columns", summary_columns[:2]) if col in summary_columns]
        if not frozen_columns:
            frozen_columns = [col for col in ("店铺", "报表类型", "年月") if col in summary_columns]
            if not frozen_columns:
                frozen_columns = summary_columns[:2] if len(summary_columns) >= 2 else ["店铺", "年月"]
        active_columns = [col for col in summary_columns if col not in frozen_columns]
        batch_size = max(100, int(config.get("page_size", 1000)))
        self.raw_edit = FieldListEditor(config["raw_columns"], "原始表格字段", batch_size=batch_size)
        self.frozen_summary_edit = FieldListEditor(frozen_columns, "冻结栏字段", batch_size=batch_size)
        self.active_summary_edit = FieldListEditor(active_columns, "活动栏字段", batch_size=batch_size)
        self.formula_edit = FormulaEditor(config["formula_note"], batch_size=batch_size)
        self.page_size_spin = QSpinBox()
        self.page_size_spin.setRange(100, 100000)
        self.page_size_spin.setSingleStep(100)
        self.page_size_spin.setValue(int(config.get("page_size", 1000)))

        tabs = QTabWidget()
        tabs.addTab(self.raw_edit, "原始表格字段")
        summary_widget = QWidget()
        summary_layout = QHBoxLayout(summary_widget)
        summary_layout.addWidget(self.frozen_summary_edit, 1)
        move_layout = QVBoxLayout()
        to_frozen_btn = QPushButton("移入冻结栏")
        to_active_btn = QPushButton("移出冻结栏")
        to_frozen_btn.clicked.connect(self.move_to_frozen)
        to_active_btn.clicked.connect(self.move_to_active)
        move_layout.addStretch()
        move_layout.addWidget(to_frozen_btn)
        move_layout.addWidget(to_active_btn)
        move_layout.addStretch()
        summary_layout.addLayout(move_layout)
        summary_layout.addWidget(self.active_summary_edit, 2)
        tabs.addTab(summary_widget, "汇总表格字段")
        tabs.addTab(self.formula_edit, "计算口径")
        page_widget = QWidget()
        page_layout = QFormLayout(page_widget)
        page_layout.addRow("明细每页行数 / 导出分批行数", self.page_size_spin)
        page_hint = QLabel("默认 1000。该数值同时用于明细分页、导出分批、备份还原分块和大数据查询进度刷新。")
        page_hint.setObjectName("mutedLabel")
        page_layout.addRow("", page_hint)
        tabs.addTab(page_widget, "分页配置")

        hint = QLabel("每行一个字段。原始字段会在导入店铺表格时自动生成初始值，后续可手工调整。")
        hint.setObjectName("mutedLabel")
        clone_row = QHBoxLayout()
        self.clone_store_combo = QComboBox()
        clone_targets = [row["name"] for row in self.repo.configured_stores() if row["name"] != self.store]
        self.clone_store_combo.addItems(clone_targets)
        clone_btn = QPushButton("复制到店铺")
        clone_btn.clicked.connect(self.clone_to_store)
        clone_row.addWidget(QLabel("参数复制"))
        clone_row.addWidget(self.clone_store_combo, 1)
        clone_row.addWidget(clone_btn)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        localize_buttons(buttons, "保存参数", "退出")
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout = QVBoxLayout(self)
        layout.addWidget(hint)
        layout.addLayout(clone_row)
        layout.addWidget(tabs, 1)
        layout.addWidget(buttons)
        self.apply_saved_column_widths()
        self.initial_state = self.snapshot_state()

    def accept(self):
        if self.save_config():
            QMessageBox.information(self, "保存成功", "参数配置已保存，可以继续编辑。")

    def reject(self):
        action = self.confirm_unsaved_close()
        if action == "save":
            super().accept()
        elif action == "discard":
            super().reject()

    def closeEvent(self, event):
        action = self.confirm_unsaved_close()
        if action == "cancel":
            event.ignore()
            return
        event.accept()
        if action == "save":
            self.done(QDialog.Accepted)

    def save_config(self):
        focus = QApplication.focusWidget()
        if focus:
            focus.clearFocus()
        raw_columns = self.raw_edit.values()
        summary_columns = self.frozen_summary_edit.values() + self.active_summary_edit.values()
        if not raw_columns or not summary_columns:
            QMessageBox.warning(self, "请补充字段", "原始字段和汇总字段不能为空。")
            return False
        frozen_columns = self.frozen_summary_edit.values()
        self.repo.save_store_config(self.store, raw_columns, summary_columns, self.formula_edit.text(), frozen_columns, self.page_size_spin.value())
        self.repo.set_app_setting(self.column_width_setting_key(), json.dumps(self.current_column_widths(), ensure_ascii=False))
        self.initial_state = self.snapshot_state()
        self.saved = True
        return True

    def confirm_unsaved_close(self):
        if self.snapshot_state() == self.initial_state:
            return "discard"
        message = QMessageBox(self)
        message.setIcon(QMessageBox.Question)
        message.setWindowTitle("是否保存参数配置")
        message.setText("参数配置已有修改，是否保存后关闭？")
        save_btn = message.addButton("保存并关闭", QMessageBox.AcceptRole)
        discard_btn = message.addButton("不保存关闭", QMessageBox.DestructiveRole)
        cancel_btn = message.addButton("继续编辑", QMessageBox.RejectRole)
        message.setDefaultButton(save_btn)
        message.exec()
        clicked = message.clickedButton()
        if clicked == save_btn:
            return "save" if self.save_config() else "cancel"
        if clicked == discard_btn:
            return "discard"
        if clicked == cancel_btn:
            return "cancel"
        return "cancel"

    def column_width_setting_key(self):
        return f"store_config_column_widths:{self.store}"

    def config_tables(self):
        return {
            "raw": self.raw_edit.table,
            "frozen": self.frozen_summary_edit.table,
            "active": self.active_summary_edit.table,
            "formula": self.formula_edit.table,
        }

    def current_column_widths(self):
        result = {}
        for name, table in self.config_tables().items():
            result[name] = [table.columnWidth(col) for col in range(table.columnCount())]
        return result

    def apply_saved_column_widths(self):
        try:
            saved = json.loads(self.repo.get_app_setting(self.column_width_setting_key(), "{}") or "{}")
        except Exception:
            saved = {}
        defaults = {
            "raw": [260],
            "frozen": [180],
            "active": [220],
            "formula": [180, 520],
        }
        for name, table in self.config_tables().items():
            widths = saved.get(name) or defaults.get(name, [])
            for col, width in enumerate(widths[:table.columnCount()]):
                table.setColumnWidth(col, max(80, int(width or 0)))

    def snapshot_state(self):
        return {
            "raw_columns": self.raw_edit.values(),
            "frozen_columns": self.frozen_summary_edit.values(),
            "active_columns": self.active_summary_edit.values(),
            "formula_note": self.formula_edit.text(),
            "page_size": self.page_size_spin.value(),
            "column_widths": self.current_column_widths(),
        }

    def clone_to_store(self):
        target = self.clone_store_combo.currentText().strip()
        if not target:
            QMessageBox.warning(self, "请选择店铺", "没有可复制的目标店铺。")
            return
        if target == self.store:
            QMessageBox.warning(self, "目标重复", "不能复制到当前店铺。")
            return
        if QMessageBox.question(self, "确认复制参数", f"确定把当前参数配置复制到店铺：{target}？\n目标店铺原参数配置会被覆盖。") != QMessageBox.Yes:
            return
        raw_columns = self.raw_edit.values()
        summary_columns = self.frozen_summary_edit.values() + self.active_summary_edit.values()
        frozen_columns = self.frozen_summary_edit.values()
        if not raw_columns or not summary_columns:
            QMessageBox.warning(self, "请补充字段", "原始字段和汇总字段不能为空。")
            return
        self.repo.save_store_config(target, raw_columns, summary_columns, self.formula_edit.text(), frozen_columns, self.page_size_spin.value())
        self.repo.set_app_setting(f"store_config_column_widths:{target}", json.dumps(self.current_column_widths(), ensure_ascii=False))
        QMessageBox.information(self, "复制完成", f"已把当前参数配置复制到店铺：{target}")

    def move_to_frozen(self):
        value = self.active_summary_edit.selected_value()
        if value:
            self.active_summary_edit.remove_value(value)
            self.frozen_summary_edit.add_value(value)

    def move_to_active(self):
        value = self.frozen_summary_edit.selected_value()
        if value:
            self.frozen_summary_edit.remove_value(value)
            self.active_summary_edit.add_value(value)


class DifferenceDialog(QDialog):
    """差异明细弹窗，集中展示核对摘要、动账分组和手工调整记录。"""

    def __init__(self, parent, repo, row):
        super().__init__(parent)
        self.setWindowTitle(f"差异明细 - {row['store']} {row['month_label']}")
        self.resize(1180, 760)

        groups, adjustments = repo.difference_groups(
            row["store"],
            months=row.get("month_sorts") or [row["month_sort"]],
            report_type=row.get("report_type", "已结算"),
        )
        title = QLabel(f"{row['store']} / {row.get('report_type', '已结算')} / {row['month_label']}    差异：{'' if row['difference'] is None else money_text(row['difference'])}")
        title.setObjectName("sectionTitle")
        reason = QLabel("差异原因：" + difference_reason(row))
        reason.setWordWrap(True)
        reason.setObjectName("mutedLabel")

        summary_table = QTableWidget(0, 2)
        summary_table.setHorizontalHeaderLabels(["项目", "金额"])
        summary_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        summary_table.verticalHeader().setVisible(False)
        summary_table.setAlternatingRowColors(True)
        summary_items = [
            ("订单实付应结", row["paid_settlement"]), ("平台补贴", row["platform_subsidy"]),
            ("商家补贴", row["merchant_subsidy"]), ("结算运费", row["freight"]),
            ("订单退款", row["refund"]), ("收入净额合计", row["income_total"]),
            ("已结算佣金", row["commission"]), ("技术服务费", row["tech_fee"]),
            ("支出金额", row["expense_amount"]), ("结算金额", row["settlement_amount"]),
            ("提现金额", row["withdraw_amount"]), ("期初金额", row["opening_balance"]),
            ("结算期末余额", row["ending_balance"]),
            ("店铺期末余额", "" if row["account_ending"] is None else row["account_ending"]),
            ("差异", "" if row["difference"] is None else row["difference"]),
        ]
        summary_table.setRowCount(len(summary_items))
        for r, (name, value) in enumerate(summary_items):
            summary_table.setItem(r, 0, table_item(name, Qt.AlignLeft | Qt.AlignVCenter))
            summary_table.setItem(r, 1, table_item("" if value == "" else money_text(value)))

        group_headers = ["方向", "摘要", "笔数", "动账金额", "订单实付应结", "平台补贴", "商家补贴", "结算运费", "订单退款", "收入净额合计", "已结算佣金", "技术服务费", "支出金额"]
        group_table = QTableWidget(len(groups), len(group_headers))
        group_table.setHorizontalHeaderLabels(group_headers)
        group_table.verticalHeader().setVisible(False)
        group_table.setAlternatingRowColors(True)
        group_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        for r, item in enumerate(groups):
            values = [
                item["direction"], item["summary"], item["count"], money_text(item["amount"]),
                money_text(item["paid_settlement"]), money_text(item["platform_subsidy"]),
                money_text(item["merchant_subsidy"]), money_text(item["freight"]),
                money_text(item["refund"]), money_text(item["income_total"]),
                money_text(item["commission"]), money_text(item["tech_fee"]), money_text(item["expense_amount"]),
            ]
            for c, value in enumerate(values):
                align = Qt.AlignLeft | Qt.AlignVCenter if c == 1 else Qt.AlignCenter
                group_table.setItem(r, c, table_item(value, align))
        group_table.resizeColumnsToContents()
        for col, width in enumerate([90, 260, 80, 120, 130, 120, 120, 120, 120, 130, 130, 120, 120]):
            group_table.setColumnWidth(col, width)

        adj_headers = ["ID", "年月", "调整列", "事项", "金额", "说明"]
        adj_table = QTableWidget(len(adjustments), len(adj_headers))
        adj_table.setHorizontalHeaderLabels(adj_headers)
        adj_table.verticalHeader().setVisible(False)
        adj_table.setAlternatingRowColors(True)
        adj_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        for r, item in enumerate(adjustments):
            values = [item["id"], item["month_sort"], item["target_column"] or "备查（不影响汇总）", item["item"], money_text(item["amount"]), item["note"] or ""]
            for c, value in enumerate(values):
                adj_table.setItem(r, c, table_item(value, Qt.AlignLeft | Qt.AlignVCenter if c in (2, 3, 5) else Qt.AlignCenter))
        adj_table.resizeColumnsToContents()
        for col, width in enumerate([70, 100, 180, 180, 110, 320]):
            adj_table.setColumnWidth(col, width)

        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.addWidget(reason)
        formula = QLabel("核对公式：差异 = 结算期末余额 - 店铺期末余额；结算期末余额 = 期初金额 + 收入净额合计 + 支出金额 + 提现金额。")
        formula.setWordWrap(True)
        formula.setObjectName("mutedLabel")
        content_layout.addWidget(formula)
        content_layout.addWidget(QLabel("核对摘要"))
        summary_table.setMinimumHeight(360)
        content_layout.addWidget(summary_table)
        content_layout.addWidget(QLabel("按动账方向/摘要汇总"))
        group_table.setMinimumHeight(300)
        content_layout.addWidget(group_table)
        content_layout.addWidget(QLabel("手工调整记录"))
        adj_table.setMinimumHeight(220)
        content_layout.addWidget(adj_table)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(content)

        layout = QVBoxLayout(self)
        layout.addWidget(title)
        layout.addWidget(scroll, 1)


class MainWindow(QMainWindow):
    """应用主窗口。

    主窗口负责把店铺、报表类型、年月区间、汇总表和明细表串起来；
    所有数据查询都通过 Repository 完成。大数据场景下只渲染当前页明细，
    并把筛选条件暂存到用户点击“开始筛选”后再执行。
    """

    def __init__(self):
        super().__init__()
        # Repository 是全应用唯一的数据入口，主窗口只保存当前 UI 状态。
        self.repo = Repository(DB_PATH)
        self.summary_rows = []
        self.current_summary_key = None
        self.detail_sort_order = "ASC"
        self.detail_sort_column = "动账时间"
        self.detail_filter_column = ""
        self.detail_filter_text = ""
        self.detail_filters = []
        self.pending_detail_filters = []
        self.detail_filter_value_cache = {}
        self.detail_page_size = 1000
        self.detail_page = 1
        self.detail_total_count = 0
        self.app_settings = self.load_app_settings()
        self.tray_icon = None
        self.force_quit = False
        self.setWindowTitle(f"{APP_TITLE} v{APP_VERSION}")
        icon_path = resource_path("assets/app_icon.ico")
        app_icon = QIcon(str(icon_path)) if icon_path.exists() else QIcon(str(sys.executable))
        if not app_icon.isNull():
            self.setWindowIcon(app_icon)
            app = QApplication.instance()
            if app:
                app.setWindowIcon(app_icon)
        self.resize(1480, 920)
        self.build_ui()
        self.apply_style()
        self.setup_tray_icon()
        self.refresh_store_combo()
        self.clear_loaded_data("请选择店铺和年月后点击“查询”。")

    def build_ui(self):
        """搭建首页布局和工具栏。"""

        toolbar = QToolBar("主操作")
        toolbar.setMovable(False)
        self.addToolBar(toolbar)
        for text, handler in [
            ("导入流水", self.import_excel),
            ("店铺配置", self.manage_stores),
            ("参数配置", self.configure_current_store),
            ("录入余额", self.edit_balance),
            ("差异调整", self.add_adjustment),
            ("差异明细", self.show_difference),
            ("导出", self.export_summary),
            ("备份及还原", self.open_backup_restore),
            ("设置", self.open_global_settings),
        ]:
            action = QAction(text, self)
            action.triggered.connect(handler)
            toolbar.addAction(action)

        root = QWidget()
        root_layout = QVBoxLayout(root)
        root_layout.setContentsMargins(14, 12, 14, 12)
        root_layout.setSpacing(10)

        title = QLabel("财务第三方支付核对")
        title.setObjectName("titleLabel")
        subtitle = QLabel("当前视图按店铺和报表类型隔离数据：先选择店铺、报表类型，再导入、查询、录入余额和差异调整。参数配置可按店铺单独维护。")
        subtitle.setObjectName("formulaLabel")
        subtitle.setWordWrap(True)
        root_layout.addWidget(title)
        root_layout.addWidget(subtitle)

        filter_bar = QFrame()
        filter_bar.setObjectName("filterBar")
        filter_layout = QHBoxLayout(filter_bar)
        self.store_combo = QComboBox()
        self.store_combo.setMinimumWidth(100)
        self.store_combo.currentTextChanged.connect(self.on_store_changed)
        self.report_type_combo = QComboBox()
        self.report_type_combo.setMinimumWidth(90)
        self.report_type_combo.addItems(["全部"] + REPORT_TYPES)
        self.month_filter = QLineEdit()
        self.month_filter.setMinimumWidth(120)
        self.month_filter.setPlaceholderText("按年月筛选，如 2026-06 或 202606-202607")
        month_pick_btn = QPushButton("选择年月")
        search_btn = QPushButton("查询")
        clear_btn = QPushButton("清空")
        month_pick_btn.clicked.connect(self.select_months)
        search_btn.clicked.connect(self.refresh_all)
        clear_btn.clicked.connect(self.clear_filters)
        filter_layout.addWidget(QLabel("店铺"))
        filter_layout.addWidget(self.store_combo, 1)
        filter_layout.addWidget(QLabel("报表类型"))
        filter_layout.addWidget(self.report_type_combo)
        filter_layout.addWidget(QLabel("年月"))
        filter_layout.addWidget(self.month_filter, 1)
        filter_layout.addWidget(month_pick_btn)
        filter_layout.addWidget(search_btn)
        filter_layout.addWidget(clear_btn)
        root_layout.addWidget(filter_bar)

        splitter = QSplitter(Qt.Vertical)
        root_layout.addWidget(splitter, 1)

        summary_widget = QWidget()
        summary_layout = QVBoxLayout(summary_widget)
        summary_layout.setContentsMargins(0, 0, 0, 0)
        summary_title = QLabel("按店铺年月汇总（冻结店铺、年月两列）")
        summary_title.setObjectName("sectionTitle")
        summary_layout.addWidget(summary_title)
        self.summary_splitter = QSplitter(Qt.Horizontal)
        default_frozen_headers = [header for header in ("店铺", "报表类型", "年月") if header in SUMMARY_COLUMNS]
        default_scroll_headers = [header for header in SUMMARY_COLUMNS + SUMMARY_EXTRA_COLUMNS if header not in default_frozen_headers]
        self.summary_fixed = QTableWidget(0, len(default_frozen_headers))
        self.summary_scroll = QTableWidget(0, len(default_scroll_headers))
        self.summary_fixed.setHorizontalHeaderLabels([display_header(header) for header in default_frozen_headers])
        self.summary_scroll.setHorizontalHeaderLabels([display_header(header) for header in default_scroll_headers])
        self.configure_table(self.summary_fixed)
        self.configure_table(self.summary_scroll)
        self.summary_fixed.setMinimumWidth(0)
        self.summary_scroll.setMinimumWidth(160)
        self.summary_fixed.verticalScrollBar().valueChanged.connect(self.summary_scroll.verticalScrollBar().setValue)
        self.summary_scroll.verticalScrollBar().valueChanged.connect(self.summary_fixed.verticalScrollBar().setValue)
        self.summary_fixed.itemSelectionChanged.connect(self.on_fixed_selection)
        self.summary_scroll.itemSelectionChanged.connect(self.on_scroll_selection)
        self.summary_fixed.cellDoubleClicked.connect(lambda _r, _c: self.edit_balance())
        self.summary_scroll.cellDoubleClicked.connect(lambda _r, _c: self.edit_balance())
        self.summary_splitter.addWidget(self.summary_fixed)
        self.summary_splitter.addWidget(self.summary_scroll)
        self.summary_splitter.setStretchFactor(0, 0)
        self.summary_splitter.setStretchFactor(1, 1)
        self.summary_splitter.setSizes([300, 1020])
        summary_layout.addWidget(self.summary_splitter, 1)

        detail_widget = QWidget()
        detail_layout = QVBoxLayout(detail_widget)
        detail_layout.setContentsMargins(0, 0, 0, 0)
        detail_title = QLabel("明细流水")
        detail_title.setObjectName("sectionTitle")
        detail_layout.addWidget(detail_title)
        self.detail_table = QTableWidget(0, len(["ID"] + RAW_COLUMNS))
        self.detail_table.setHorizontalHeaderLabels(["ID"] + RAW_COLUMNS)
        self.configure_table(self.detail_table)
        self.detail_table.horizontalHeader().sectionClicked.connect(self.sort_details_by_header)
        self.detail_splitter = QSplitter(Qt.Vertical)
        self.detail_splitter.addWidget(self.detail_table)
        detail_controls = QWidget()
        detail_controls_layout = QVBoxLayout(detail_controls)
        detail_controls_layout.setContentsMargins(0, 6, 0, 0)
        detail_controls_layout.setSpacing(4)
        detail_filter_bar = QHBoxLayout()
        detail_status_bar = QHBoxLayout()
        detail_action_bar = QHBoxLayout()
        detail_filter_bar.setSpacing(4)
        detail_status_bar.setSpacing(4)
        detail_action_bar.setSpacing(4)
        edit_tx = QPushButton("编辑选中流水")
        del_tx = QPushButton("删除选中流水")
        del_month = QPushButton("删除选中年月流水")
        edit_tx.clicked.connect(self.edit_transaction)
        del_tx.clicked.connect(self.delete_transaction)
        del_month.clicked.connect(self.delete_month)
        self.detail_filter_label = QLabel("明细：点击表头排序")
        self.detail_filter_label.setObjectName("mutedLabel")
        self.detail_active_filters_label = QLabel("已应用：无；待应用：无")
        self.detail_active_filters_label.setObjectName("mutedLabel")
        self.detail_active_filters_label.setWordWrap(True)
        self.detail_active_filters_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
        self.detail_filter_hint = QLabel("可连续加入多个条件，确认后点“开始筛选”。")
        self.detail_filter_hint.setObjectName("mutedLabel")
        self.detail_filter_column_combo = QComboBox()
        self.detail_filter_column_combo.setMinimumWidth(120)
        self.detail_filter_column_combo.setMinimumContentsLength(14)
        self.detail_filter_column_combo.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self.detail_filter_column_combo.currentTextChanged.connect(self.refresh_detail_filter_values)
        self.detail_filter_value_combo = QComboBox()
        self.detail_filter_value_combo.setMinimumWidth(140)
        self.detail_filter_value_combo.setMinimumContentsLength(18)
        self.detail_filter_value_combo.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        add_filter_btn = QPushButton("加入添加筛选条件")
        add_filter_btn.clicked.connect(self.add_pending_detail_filter)
        run_filter_btn = QPushButton("开始筛选")
        run_filter_btn.clicked.connect(self.apply_detail_filter)
        clear_filter_btn = QPushButton("清除筛选条件")
        clear_filter_btn.clicked.connect(self.clear_detail_filter)
        first_page_btn = QPushButton("首页")
        prev_page_btn = QPushButton("上一页")
        next_page_btn = QPushButton("下一页")
        last_page_btn = QPushButton("末页")
        self.detail_page_edit = QLineEdit("1")
        self.detail_page_edit.setFixedWidth(56)
        page_jump_btn = QPushButton("跳页")
        self.detail_page_label = QLabel("第 1/1 页")
        self.detail_page_label.setObjectName("mutedLabel")
        first_page_btn.clicked.connect(lambda: self.go_detail_page("first"))
        prev_page_btn.clicked.connect(lambda: self.go_detail_page("prev"))
        next_page_btn.clicked.connect(lambda: self.go_detail_page("next"))
        last_page_btn.clicked.connect(lambda: self.go_detail_page("last"))
        page_jump_btn.clicked.connect(lambda: self.go_detail_page("jump"))
        detail_filter_bar.addWidget(QLabel("筛选字段"))
        detail_filter_bar.addWidget(self.detail_filter_column_combo)
        detail_filter_bar.addWidget(QLabel("筛选值"))
        detail_filter_bar.addWidget(self.detail_filter_value_combo, 1)
        detail_filter_bar.addWidget(add_filter_btn)
        detail_filter_bar.addWidget(run_filter_btn)
        detail_filter_bar.addWidget(clear_filter_btn)
        detail_filter_bar.addWidget(self.detail_filter_hint)
        detail_filter_bar.addStretch()
        detail_status_bar.addWidget(self.detail_active_filters_label, 1)
        detail_status_bar.addWidget(self.detail_filter_label)
        detail_action_bar.addWidget(first_page_btn)
        detail_action_bar.addWidget(prev_page_btn)
        detail_action_bar.addWidget(next_page_btn)
        detail_action_bar.addWidget(last_page_btn)
        detail_action_bar.addWidget(QLabel("页"))
        detail_action_bar.addWidget(self.detail_page_edit)
        detail_action_bar.addWidget(page_jump_btn)
        detail_action_bar.addWidget(self.detail_page_label)
        detail_action_bar.addSpacing(6)
        detail_action_bar.addWidget(edit_tx)
        detail_action_bar.addWidget(del_tx)
        detail_action_bar.addWidget(del_month)
        detail_action_bar.addStretch()
        detail_controls_layout.addLayout(detail_filter_bar)
        detail_controls_layout.addLayout(detail_status_bar)
        detail_controls_layout.addLayout(detail_action_bar)
        self.detail_splitter.addWidget(detail_controls)
        self.detail_splitter.setSizes([520, 132])
        self.detail_splitter.setChildrenCollapsible(False)
        detail_layout.addWidget(self.detail_splitter, 1)

        splitter.addWidget(summary_widget)
        splitter.addWidget(detail_widget)
        splitter.setSizes([390, 430])

        self.status = QLabel("")
        self.status.setObjectName("statusLabel")
        self.status.setWordWrap(False)
        self.status.setMinimumWidth(0)
        self.status.setMaximumWidth(920)
        self.status.setSizePolicy(QSizePolicy.Maximum, QSizePolicy.Fixed)
        status_row = QHBoxLayout()
        status_row.addWidget(self.status)
        status_row.addStretch()
        root_layout.addLayout(status_row)
        self.setCentralWidget(root)

    def configure_table(self, table):
        """统一表格样式，关闭自动换行以降低大数据渲染成本。"""

        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setSelectionMode(QTableWidget.SingleSelection)
        table.verticalHeader().setVisible(False)
        table.setWordWrap(False)
        table.horizontalHeader().setStretchLastSection(False)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)

    def set_status(self, text, tooltip=None):
        """设置底部状态提示，长文本在界面上省略、完整内容放到 tooltip。"""

        text = str(text or "")
        tooltip = text if tooltip is None else str(tooltip or "")
        metrics = QFontMetrics(self.status.font())
        max_width = max(220, min(920, self.width() - 80))
        shown = metrics.elidedText(text, Qt.ElideMiddle, max_width - 28)
        self.status.setMaximumWidth(max_width)
        self.status.setText(shown)
        self.status.setToolTip(tooltip)

    def load_app_settings(self):
        return {
            "startup_enabled": self.repo.get_app_setting("startup_enabled", "0") == "1",
            "close_to_tray": self.repo.get_app_setting("close_to_tray", "0") == "1",
        }

    def save_app_settings(self):
        self.repo.set_app_setting("startup_enabled", "1" if self.app_settings.get("startup_enabled") else "0")
        self.repo.set_app_setting("close_to_tray", "1" if self.app_settings.get("close_to_tray") else "0")

    def setup_tray_icon(self):
        if not QSystemTrayIcon.isSystemTrayAvailable():
            return
        icon = self.windowIcon()
        if icon.isNull():
            icon = QIcon(str(resource_path("assets/app_icon.ico")))
        self.tray_icon = QSystemTrayIcon(icon, self)
        self.tray_icon.setToolTip(APP_TITLE)
        menu = QMenu(self)
        show_action = menu.addAction("显示主窗口")
        show_action.triggered.connect(self.restore_from_tray)
        quit_action = menu.addAction("退出程序")
        quit_action.triggered.connect(self.quit_from_tray)
        self.tray_icon.setContextMenu(menu)
        self.tray_icon.activated.connect(self.on_tray_activated)
        self.tray_icon.show()

    def on_tray_activated(self, reason):
        if reason in (QSystemTrayIcon.Trigger, QSystemTrayIcon.DoubleClick):
            self.restore_from_tray()

    def restore_from_tray(self):
        self.show()
        self.raise_()
        self.activateWindow()

    def quit_from_tray(self):
        self.force_quit = True
        QApplication.quit()

    def closeEvent(self, event):
        if self.app_settings.get("close_to_tray") and not self.force_quit and self.tray_icon:
            event.ignore()
            self.hide()
            self.tray_icon.showMessage(APP_TITLE, "程序已最小化到托盘。右键托盘图标可退出。", QSystemTrayIcon.Information, 1800)
            return
        self.force_quit = True
        event.accept()

    def startup_command(self):
        if getattr(sys, "frozen", False):
            return f'"{sys.executable}"'
        return f'"{sys.executable}" "{Path(__file__).resolve()}"'

    def set_startup_enabled(self, enabled):
        if sys.platform != "win32":
            return
        import winreg
        key_path = r"Software\Microsoft\Windows\CurrentVersion\Run"
        app_key = "PaymentReconciliationQt"
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_path, 0, winreg.KEY_SET_VALUE) as key:
            if enabled:
                winreg.SetValueEx(key, app_key, 0, winreg.REG_SZ, self.startup_command())
            else:
                try:
                    winreg.DeleteValue(key, app_key)
                except FileNotFoundError:
                    pass

    def open_global_settings(self):
        dialog = GlobalSettingsDialog(self, self.app_settings)
        if dialog.exec() != QDialog.Accepted:
            return
        old_startup = self.app_settings.get("startup_enabled")
        self.app_settings.update(dialog.result_data)
        try:
            if old_startup != self.app_settings.get("startup_enabled"):
                self.set_startup_enabled(self.app_settings.get("startup_enabled"))
            self.save_app_settings()
            self.set_status("全局设置已保存")
            QMessageBox.information(self, "保存成功", "全局设置已保存。")
        except Exception as exc:
            QMessageBox.warning(self, "保存失败", f"全局设置保存失败：{exc}")

    def open_backup_restore(self):
        dialog = BackupRestoreDialog(self, self.repo)
        if dialog.exec() == QDialog.Accepted:
            self.app_settings = self.load_app_settings()
            self.refresh_store_combo()
            self.clear_loaded_data("备份还原已完成。点击“查询”重新加载当前店铺数据。")

    def apply_style(self):
        self.setStyleSheet("""
            QMainWindow, QWidget { background: #f5f7fb; color: #1f2937; font-family: "Microsoft YaHei"; font-size: 12px; }
            QToolBar { background: #ffffff; border-bottom: 1px solid #d9e2ef; spacing: 2px; padding: 4px; }
            QToolButton { background: #2563eb; color: #ffffff; border: none; border-radius: 6px; padding: 5px 8px; margin: 0px; }
            QToolButton:hover { background: #1d4ed8; }
            QPushButton { background: #ffffff; border: 1px solid #cbd5e1; border-radius: 6px; padding: 5px 8px; margin: 0px; }
            QPushButton:hover { background: #eff6ff; border-color: #93c5fd; }
            QLineEdit, QComboBox { background: #ffffff; border: 1px solid #b8c3d4; border-radius: 6px; padding: 7px 12px; min-height: 20px; }
            QComboBox { padding-right: 34px; }
            QComboBox:hover { border-color: #3b82f6; background: #f8fbff; }
            QComboBox:focus { border: 1px solid #2563eb; background: #ffffff; }
            QComboBox::drop-down { subcontrol-origin: padding; subcontrol-position: top right; width: 30px; border-left: 1px solid #d7deea; background: #f1f5fb; border-top-right-radius: 6px; border-bottom-right-radius: 6px; }
            QComboBox::down-arrow { width: 10px; height: 10px; }
            QComboBox QAbstractItemView { background: #ffffff; color: #172033; border: 1px solid #8fa1b8; border-radius: 6px; selection-background-color: #dbeafe; selection-color: #0f172a; padding: 6px; outline: 0; }
            QComboBox QAbstractItemView::item { min-height: 28px; padding: 6px 10px; }
            QTableWidget { background: #ffffff; alternate-background-color: #f3f6fb; color: #172033; gridline-color: #cbd5e1; border: 1px solid #cbd5e1; border-radius: 6px; selection-background-color: #dbeafe; selection-color: #0f172a; }
            QHeaderView::section { background: #dbe7f6; color: #172033; padding: 7px; border: 0; border-right: 1px solid #b8c7dc; font-weight: 700; }
            QScrollBar:vertical, QScrollBar:horizontal { background: #e2e8f0; border-radius: 6px; margin: 0px; }
            QScrollBar::handle:vertical, QScrollBar::handle:horizontal { background: #64748b; border-radius: 6px; min-height: 28px; min-width: 28px; }
            QScrollBar::handle:vertical:hover, QScrollBar::handle:horizontal:hover { background: #334155; }
            QScrollBar::add-line, QScrollBar::sub-line { width: 0px; height: 0px; }
            QLabel#titleLabel { font-size: 24px; font-weight: 700; color: #0f172a; }
            QLabel#subtitleLabel, QLabel#formulaLabel, QLabel#mutedLabel { color: #64748b; }
            QLabel#statusLabel { color: #0f3d66; background: #eef7ff; border: 1px solid #bfdbfe; border-radius: 10px; padding: 6px 12px; font-weight: 600; }
            QLabel#sectionTitle { font-size: 14px; font-weight: 700; color: #0f172a; padding: 4px 0; }
            QFrame#filterBar { background: #ffffff; border: 1px solid #d9e2ef; border-radius: 8px; }
        """)

    def clear_filters(self):
        self.month_filter.clear()
        self.refresh_all()

    def reset_summary_headers(self):
        headers = self.current_summary_columns()
        frozen_headers = self.current_frozen_columns(headers)
        scroll_headers = [header for header in headers if header not in frozen_headers]
        self.summary_fixed.setColumnCount(len(frozen_headers))
        self.summary_scroll.setColumnCount(len(scroll_headers))
        self.summary_fixed.setHorizontalHeaderLabels([display_header(header) for header in frozen_headers])
        self.summary_scroll.setHorizontalHeaderLabels([display_header(header) for header in scroll_headers])

    def clear_loaded_data(self, message=""):
        self.summary_rows = []
        self.current_summary_key = None
        self.reset_summary_headers()
        self.summary_fixed.setRowCount(0)
        self.summary_scroll.setRowCount(0)
        self.detail_table.setRowCount(0)
        self.set_status(message)

    def refresh_store_combo(self):
        current = self.current_store()
        stores = [row["name"] for row in self.repo.configured_stores()]
        self.store_combo.blockSignals(True)
        self.store_combo.clear()
        self.store_combo.addItems(stores)
        if current in stores:
            self.store_combo.setCurrentText(current)
        self.store_combo.blockSignals(False)
        if self.current_store():
            self.repo.use_store(self.current_store())

    def current_store(self):
        return self.store_combo.currentText().strip() if hasattr(self, "store_combo") else ""

    def on_store_changed(self, store):
        if store.strip():
            self.repo.use_store(store.strip())
        self.current_summary_key = None
        self.clear_loaded_data(f"已切换到店铺：{store.strip()}。点击“查询”加载汇总和明细。")

    def select_months(self):
        if not self.current_store():
            QMessageBox.warning(self, "未选择店铺", "请先选择店铺。")
            return
        dialog = MonthSelectDialog(self, self.repo.months(), parse_month_filter(self.month_filter.text()))
        if dialog.exec() == QDialog.Accepted:
            self.month_filter.setText(",".join(dialog.result_months or []))
            self.refresh_all()

    def current_report_type_filter(self):
        value = self.report_type_combo.currentText().strip() if hasattr(self, "report_type_combo") else ""
        return "" if value == "全部" else value

    def current_page_size(self):
        store = self.current_store()
        if not store:
            return 1000
        try:
            return max(100, int(self.repo.store_config(store).get("page_size", 1000)))
        except Exception:
            return 1000

    def sort_details_by_header(self, index):
        header = self.detail_table.horizontalHeaderItem(index)
        if not header:
            return
        column = header.text()
        self.detail_sort_order = "DESC" if self.detail_sort_column == column and self.detail_sort_order == "ASC" else "ASC"
        self.detail_sort_column = column
        self.detail_page = 1
        row = self.selected_summary()
        if row:
            self.refresh_details(row, refresh_filter_values=False)

    def clear_detail_filter(self):
        self.detail_filter_column = ""
        self.detail_filter_text = ""
        self.detail_filters = []
        self.pending_detail_filters = []
        self.detail_page = 1
        self.detail_filter_value_combo.clear()
        self.update_detail_filter_labels()
        row = self.selected_summary()
        if row:
            self.refresh_details(row, refresh_filter_values=True)

    def refresh_all(self):
        if not self.current_store():
            self.clear_loaded_data("请先在店铺配置中新增店铺。")
            return
        self.set_status("正在加载汇总数据...")
        QApplication.processEvents()
        self.refresh_summary()
        QApplication.processEvents()
        if self.summary_rows:
            self.summary_fixed.selectRow(0)
            self.summary_scroll.selectRow(0)
            self.select_summary_row(0)
        else:
            self.refresh_details(None)

    def refresh_summary(self):
        """按当前店铺、报表类型、年月区间刷新汇总表。"""

        store = self.current_store()
        report_type = self.current_report_type_filter()
        self.detail_filter_value_cache.clear()
        self.summary_rows = self.repo.monthly_summaries(
            store,
            self.month_filter.text().strip(),
            report_type_filter=report_type,
            progress_callback=lambda message, value, total: self.set_status(message),
        )
        headers = self.current_summary_columns()
        frozen_headers = self.current_frozen_columns(headers)
        scroll_headers = [header for header in headers if header not in frozen_headers]
        self.summary_fixed.setColumnCount(len(frozen_headers))
        self.summary_scroll.setColumnCount(len(scroll_headers))
        self.summary_fixed.setHorizontalHeaderLabels([display_header(header) for header in frozen_headers])
        self.summary_scroll.setHorizontalHeaderLabels([display_header(header) for header in scroll_headers])
        self.summary_fixed.setRowCount(len(self.summary_rows))
        self.summary_scroll.setRowCount(len(self.summary_rows))
        self.summary_fixed.setUpdatesEnabled(False)
        self.summary_scroll.setUpdatesEnabled(False)
        for row_idx, row in enumerate(self.summary_rows):
            if row_idx and row_idx % max(1, self.current_page_size()) == 0:
                self.set_status(f"正在填充汇总表：{row_idx}/{len(self.summary_rows)} 条")
                QApplication.processEvents()
            fixed_values = [self.summary_value(row, header) for header in frozen_headers]
            scroll_values = [self.summary_value(row, header) for header in scroll_headers]
            has_diff = row["difference"] is not None and row["difference"] != Decimal("0.00")
            bg = QColor("#fee2e2") if has_diff else (QColor("#eaf4ff") if row.get("is_aggregate") else None)
            for col, value in enumerate(fixed_values):
                item = table_item(value)
                if bg:
                    item.setBackground(bg)
                    item.setForeground(QColor("#111827"))
                self.summary_fixed.setItem(row_idx, col, item)
            for col, value in enumerate(scroll_values):
                item = table_item(value, Qt.AlignLeft | Qt.AlignVCenter if col == len(scroll_values) - 1 else Qt.AlignCenter)
                if bg:
                    item.setBackground(bg)
                    item.setForeground(QColor("#111827"))
                self.summary_scroll.setItem(row_idx, col, item)
        self.summary_fixed.setUpdatesEnabled(True)
        self.summary_scroll.setUpdatesEnabled(True)
        self.auto_fit(self.summary_fixed, max_width=180)
        self.auto_fit(self.summary_scroll, max_width=260)
        fixed_width = sum(self.summary_fixed.columnWidth(col) for col in range(self.summary_fixed.columnCount())) + 24
        current_sizes = self.summary_splitter.sizes() if hasattr(self, "summary_splitter") else []
        if not current_sizes or sum(current_sizes) == 0:
            self.summary_splitter.setSizes([min(fixed_width, 520), 1000])
        type_text = report_type or "全部"
        db_path = Path(self.repo.store_db_path(store))
        self.set_status(
            f"当前店铺：{store}    报表类型：{type_text}    数据库：{db_path.name}    汇总 {len(self.summary_rows)} 条",
            f"店铺数据库：{db_path}",
        )

    def current_summary_columns(self):
        store = self.current_store()
        if store:
            columns = self.repo.store_config(store)["summary_columns"]
        else:
            columns = SUMMARY_COLUMNS + SUMMARY_EXTRA_COLUMNS
        if "店铺" not in columns:
            columns.insert(0, "店铺")
        if "年月" not in columns:
            insert_at = 2 if "报表类型" in columns else 1
            columns.insert(insert_at, "年月")
        return columns

    def current_frozen_columns(self, headers=None):
        store = self.current_store()
        headers = headers or self.current_summary_columns()
        if store:
            config = self.repo.store_config(store)
            configured = config.get("frozen_columns") or ["店铺", "报表类型", "年月"]
            frozen = [col for col in configured if col in headers]
            if "年月" in headers and "年月" not in frozen:
                insert_at = 2 if "报表类型" in frozen else len(frozen)
                frozen.insert(insert_at, "年月")
        else:
            frozen = [col for col in ("店铺", "报表类型", "年月") if col in headers]
        return frozen or headers[:3]
    def summary_value(self, row, header):
        values = {
            "店铺": row["store"],
            "报表类型": row.get("report_type", "已结算"),
            "月份": row["month_label"],
            "年月": row["month_label"],
            "订单实付应结": money_text(row["paid_settlement"]),
            "平台补贴": money_text(row["platform_subsidy"]),
            "商家补贴": money_text(row["merchant_subsidy"]),
            "结算运费": money_text(row["freight"]),
            "订单退款": money_text(row["refund"]),
            "收入净额合计": money_text(row["income_total"]),
            "已结算佣金": money_text(row["commission"]),
            "技术服务费": money_text(row["tech_fee"]),
            "支出金额": money_text(row["expense_amount"]),
            "结算金额": money_text(row["settlement_amount"]),
            "提现金额": money_text(row["withdraw_amount"]),
            "期初金额": money_text(row["opening_balance"]),
            "结算期末余额": money_text(row["ending_balance"]),
            "店铺期末余额": "" if row["account_ending"] is None else money_text(row["account_ending"]),
            "差异": "" if row["difference"] is None else money_text(row["difference"]),
            "差异原因": difference_reason(row),
            "调整说明": row.get("adjustment_notes", ""),
        }
        if header in values:
            return values[header]
        custom_values = row.get("custom_values", {}) or {}
        if header in custom_values:
            return money_text(custom_values[header])
        return ""

    def auto_fit(self, table, max_width=280):
        table.resizeColumnsToContents()
        for col in range(table.columnCount()):
            width = max(72, min(table.columnWidth(col) + 20, max_width))
            table.setColumnWidth(col, width)

    def on_fixed_selection(self):
        row = self.summary_fixed.currentRow()
        if row >= 0 and self.summary_scroll.currentRow() != row:
            self.summary_scroll.selectRow(row)
        self.select_summary_row(row)

    def on_scroll_selection(self):
        row = self.summary_scroll.currentRow()
        if row >= 0 and self.summary_fixed.currentRow() != row:
            self.summary_fixed.selectRow(row)
        self.select_summary_row(row)

    def select_summary_row(self, row):
        if row < 0 or row >= len(self.summary_rows):
            self.current_summary_key = None
            self.refresh_details(None)
            return
        selected = self.summary_rows[row]
        key = (selected["store"], selected.get("report_type", "已结算"), selected["month_sort"], tuple(selected.get("month_sorts", [])))
        if key != self.current_summary_key:
            self.current_summary_key = key
            self.detail_page = 1
            self.pending_detail_filters = list(self.detail_filters)
            self.update_detail_filter_labels()
            self.refresh_details(selected, refresh_filter_values=True)

    def selected_summary(self):
        row = self.summary_scroll.currentRow()
        if row < 0:
            row = self.summary_fixed.currentRow()
        if 0 <= row < len(self.summary_rows):
            return self.summary_rows[row]
        return None

    def refresh_details(self, row, refresh_filter_values=False):
        """按选中的汇总行分页刷新明细流水。"""

        self.detail_table.setUpdatesEnabled(False)
        self.detail_table.setSortingEnabled(False)
        self.detail_table.setRowCount(0)
        store = row["store"] if row else self.current_store()
        raw_columns = self.repo.store_config(store)["raw_columns"] if store else RAW_COLUMNS
        headers = ["ID"] + raw_columns
        display_headers = [display_header(header) for header in headers]
        self.detail_table.setColumnCount(len(headers))
        self.detail_table.setHorizontalHeaderLabels(display_headers)
        current_filter_col = self.detail_filter_column_combo.currentText() if self.detail_filter_column_combo.count() else ""
        self.detail_filter_column_combo.blockSignals(True)
        self.detail_filter_column_combo.clear()
        self.detail_filter_column_combo.addItems(display_headers)
        if current_filter_col in display_headers:
            self.detail_filter_column_combo.setCurrentText(current_filter_col)
        self.detail_filter_column_combo.blockSignals(False)
        if refresh_filter_values:
            self.refresh_detail_filter_values()
        if not row:
            self.detail_table.setUpdatesEnabled(True)
            self.detail_page_label.setText("第 1/1 页")
            self.detail_page_edit.setText("1")
            self.update_detail_filter_labels()
            return
        months = row.get("month_sorts") or [row["month_sort"]]
        report_type = row.get("report_type", "已结算")
        self.detail_page_size = self.current_page_size()
        total_count = self.repo.details_count(
            row["store"], months=months, report_type_filter=report_type,
            filters=self.detail_filters,
        )
        self.detail_total_count = total_count
        total_pages = max(1, (total_count + self.detail_page_size - 1) // self.detail_page_size)
        self.detail_page = max(1, min(self.detail_page, total_pages))
        offset = (self.detail_page - 1) * self.detail_page_size
        rows = self.repo.details(
            row["store"],
            months=months,
            sort_order=self.detail_sort_order,
            sort_column=self.detail_sort_column,
            report_type_filter=report_type,
            limit=self.detail_page_size,
            offset=offset,
            filters=self.detail_filters,
        )
        self.detail_table.setRowCount(len(rows))
        for r, tx in enumerate(rows):
            if r and r % 500 == 0:
                self.set_status(f"正在填充明细流水：{r}/{len(rows)} 条")
                QApplication.processEvents()
            try:
                payload = json.loads(tx["raw_payload"] or "{}")
            except Exception:
                payload = {}
            fallback = {
                "店铺": tx["store"],
                "报表类型": tx["report_type"],
                "月份": tx["month_label"],
                "年月": tx["month_label"],
                "动账时间": tx["transaction_time"],
                "动账流水号": tx["flow_id"],
                "动账方向": tx["direction"],
                "动账账户": tx["account"],
                "动账金额": money_text(tx["amount"]),
                "动账摘要": tx["summary"],
                "业务类型": tx["biz_type"],
                "主订单编号": tx["main_order"],
                "子订单编号": tx["sub_order"],
                "售后单号": tx["after_sale"],
                "下单时间": tx["order_time"],
                "商品信息": tx["product_info"],
                "商品编码": tx["product_code"],
                "售卖类型": tx["sale_type"],
                "订单实付应结": money_text(tx["paid_settlement"]),
                "平台补贴": money_text(tx["platform_subsidy"]),
                "商家补贴": money_text(tx["merchant_subsidy"]),
                "结算运费": money_text(tx["freight"]),
                "订单退款": money_text(tx["refund"]),
                "佣金": money_text(tx["commission"]),
                "技术服务费": money_text(tx["tech_fee"]),
            }
            values = [tx["id"]] + [
                self.display_cell_value(display_header(header), payload.get(header, fallback.get(header, "")))
                for header in raw_columns
            ]
            for c, value in enumerate(values):
                align = Qt.AlignLeft | Qt.AlignVCenter if c > 0 and headers[c] in ("动账摘要", "商品信息") else Qt.AlignCenter
                self.detail_table.setItem(r, c, table_item(value, align))
        self.auto_fit_detail(display_headers)
        self.detail_table.setSortingEnabled(True)
        self.detail_table.setUpdatesEnabled(True)
        filter_text = "；".join(f"{c} = {v}" for c, v in self.detail_filters)
        shown = len(rows)
        start_no = offset + 1 if shown else 0
        end_no = offset + shown
        more_text = f"；总 {total_count} 条，当前 {start_no}-{end_no} 条"
        self.update_detail_filter_labels()
        self.detail_filter_label.setText(f"明细排序：{self.detail_sort_column} {'降序' if self.detail_sort_order == 'DESC' else '升序'}{more_text}")
        self.detail_page_label.setText(f"第 {self.detail_page}/{total_pages} 页")
        self.detail_page_edit.setText(str(self.detail_page))

    def display_cell_value(self, header, value):
        if value is None:
            return ""
        text = clean_cell_text(value)
        if text.lower() == "none":
            return ""
        return text

    def auto_fit_detail(self, headers):
        for col, header in enumerate(headers):
            if header in ("ID", "月份", "年月", "报表类型", "动账方向"):
                width = 90
            elif header in ("动账时间", "动账金额", "业务类型"):
                width = 140
            elif header in ("动账流水号", "主订单编号", "子订单编号", "售后单号"):
                width = 220
            elif header in ("动账摘要", "商品信息"):
                width = 260
            else:
                width = max(100, min(len(str(header)) * 14 + 36, 220))
            self.detail_table.setColumnWidth(col, width)

    def update_detail_filter_labels(self):
        """刷新明细筛选条件说明，条件过多时允许换行显示。"""

        applied = "；".join(f"{c} = {v}" for c, v in self.detail_filters) or "无"
        pending = "；".join(f"{c} = {v}" for c, v in self.pending_detail_filters) or "无"
        self.detail_active_filters_label.setText(f"已应用：{applied}\n待应用：{pending}")
        self.detail_active_filters_label.setToolTip(f"已应用：{applied}\n待应用：{pending}")

    def add_pending_detail_filter(self):
        column = self.detail_filter_column_combo.currentText().strip()
        value = self.detail_filter_value_combo.currentText().strip()
        if column and value:
            self.pending_detail_filters = [(c, v) for c, v in self.pending_detail_filters if c != column]
            self.pending_detail_filters.append((column, value))
            self.detail_filter_column = column
            self.detail_filter_text = value
            self.update_detail_filter_labels()
            self.set_status("筛选条件已加入，点击“开始筛选”后生效。")

    def apply_detail_filter(self):
        self.detail_filters = list(self.pending_detail_filters)
        self.detail_page = 1
        self.update_detail_filter_labels()
        row = self.selected_summary()
        if row:
            self.refresh_details(row, refresh_filter_values=False)

    def refresh_detail_filter_values(self):
        """加载当前筛选字段的可选值，并使用缓存减少大表重复 distinct 查询。"""

        row = self.selected_summary()
        column = self.detail_filter_column_combo.currentText().strip() if self.detail_filter_column_combo.count() else ""
        current = self.detail_filter_value_combo.currentText() if self.detail_filter_value_combo.count() else ""
        previous_status = self.status.text() if hasattr(self, "status") else ""
        previous_tooltip = self.status.toolTip() if hasattr(self, "status") else ""
        self.detail_filter_value_combo.blockSignals(True)
        self.detail_filter_value_combo.clear()
        if row and column:
            months = row.get("month_sorts") or [row["month_sort"]]
            query_column = data_header(column)
            query_filters = [(data_header(c), v) for c, v in self.detail_filters]
            cache_key = (
                row["store"],
                tuple(months),
                row.get("report_type", "已结算"),
                query_column,
                tuple((c, v) for c, v in query_filters if c != query_column),
            )
            if cache_key in self.detail_filter_value_cache:
                values = self.detail_filter_value_cache[cache_key]
            else:
                self.set_status(f"正在加载筛选值：{column} ...")
                QApplication.processEvents()
                values = self.repo.detail_distinct_values(
                    row["store"], months, row.get("report_type", "已结算"), query_column, filters=query_filters
                )
                self.detail_filter_value_cache[cache_key] = values
            self.detail_filter_value_combo.addItems(values)
            if current in values:
                self.detail_filter_value_combo.setCurrentText(current)
        self.detail_filter_value_combo.blockSignals(False)
        if previous_status:
            self.set_status(previous_status, previous_tooltip)

    def go_detail_page(self, action):
        row = self.selected_summary()
        if not row:
            return
        total_pages = max(1, (self.detail_total_count + self.detail_page_size - 1) // self.detail_page_size)
        if action == "first":
            self.detail_page = 1
        elif action == "prev":
            self.detail_page = max(1, self.detail_page - 1)
        elif action == "next":
            self.detail_page = min(total_pages, self.detail_page + 1)
        elif action == "last":
            self.detail_page = total_pages
        elif action == "jump":
            try:
                self.detail_page = max(1, min(total_pages, int(self.detail_page_edit.text().strip())))
            except ValueError:
                self.detail_page = 1
        self.refresh_details(row, refresh_filter_values=False)

    def detail_cell_value(self, tx, header):
        if header == "ID":
            return tx["id"]
        try:
            payload = json.loads(tx["raw_payload"] or "{}")
        except Exception:
            payload = {}
        fallback = {
            "店铺": tx["store"], "月份": tx["month_label"], "年月": tx["month_label"], "动账时间": tx["transaction_time"],
            "报表类型": tx["report_type"],
            "动账流水号": tx["flow_id"], "动账方向": tx["direction"], "动账账户": tx["account"],
            "动账金额": money_text(tx["amount"]), "动账摘要": tx["summary"], "业务类型": tx["biz_type"],
            "主订单编号": tx["main_order"], "子订单编号": tx["sub_order"], "售后单号": tx["after_sale"],
            "下单时间": tx["order_time"], "商品信息": tx["product_info"], "商品编码": tx["product_code"],
            "售卖类型": tx["sale_type"], "订单实付应结": money_text(tx["paid_settlement"]),
            "平台补贴": money_text(tx["platform_subsidy"]), "商家补贴": money_text(tx["merchant_subsidy"]),
            "结算运费": money_text(tx["freight"]), "订单退款": money_text(tx["refund"]),
            "佣金": money_text(tx["commission"]), "技术服务费": money_text(tx["tech_fee"]),
        }
        return payload.get(header, fallback.get(header, ""))

    def manage_stores(self):
        dialog = StoreDialog(self, self.repo)
        dialog.exec()
        self.refresh_store_combo()
        self.clear_loaded_data("店铺配置已更新。点击“查询”重新加载当前店铺数据。")

    def configure_current_store(self):
        store = self.current_store()
        if not store:
            QMessageBox.warning(self, "未选择店铺", "请先选择店铺。")
            return
        dialog = StoreConfigDialog(self, self.repo, store)
        result = dialog.exec()
        if result == QDialog.Accepted or dialog.saved:
            self.clear_loaded_data("参数配置已保存。点击“查询”重新加载当前店铺数据。")

    def import_excel(self):
        """导入 Excel/CSV 流水文件，并在状态栏显示导入进度和最终行数。"""

        stores = [row["name"] for row in self.repo.configured_stores()]
        if not stores:
            if QMessageBox.question(self, "需要配置店铺", "还没有配置店铺。是否现在添加店铺？") == QMessageBox.Yes:
                self.manage_stores()
            return
        options_dialog = ImportOptionsDialog(self, stores, self.current_store())
        if options_dialog.exec() != QDialog.Accepted or not options_dialog.result_data:
            return
        store = options_dialog.result_data["store"]
        report_type = options_dialog.result_data["report_type"]
        import_month = options_dialog.result_data["import_month"]
        if not store or not report_type or not import_month:
            return
        paths, _ = QFileDialog.getOpenFileNames(self, "选择资金流水文件", "", "流水文件 (*.xlsx *.xlsm *.csv);;Excel 文件 (*.xlsx *.xlsm);;CSV 文件 (*.csv);;所有文件 (*.*)")
        if not paths:
            return
        try:
            total_imported = 0
            total_skipped = 0
            for file_index, path in enumerate(paths):
                def update_progress(message, _value, _maximum):
                    self.set_status(message)
                    QApplication.processEvents()

                self.set_status(f"正在导入：{Path(path).name}")
                QApplication.processEvents()
                result = self.repo.import_excel(
                    path,
                    selected_store=store,
                    report_type=report_type,
                    selected_month=import_month,
                    progress_callback=update_progress,
                )
                total_imported += result.imported
                total_skipped += result.skipped
                self.set_status(f"已完成：{Path(path).name}")
                QApplication.processEvents()
            self.store_combo.setCurrentText(store)
            self.report_type_combo.setCurrentText(report_type)
            self.refresh_all()
            self.set_status(
                f"导入完成：店铺 {store}，报表类型 {report_type}，归属年月 {import_month}，新增 {total_imported} 条，跳过 {total_skipped} 条"
            )
            QMessageBox.information(self, "导入完成", f"店铺：{store}\n报表类型：{report_type}\n导入年月：{import_month}\n新增 {total_imported} 条，跳过重复/空行 {total_skipped} 条。")
        except Exception as exc:
            QMessageBox.critical(self, "导入失败", f"{exc}\n\n{traceback.format_exc(limit=2)}")

    def edit_balance(self):
        row = self.selected_summary()
        initial = {}
        if row:
            bal = self.repo.balance_for(row["store"], row["month_sort"], row.get("report_type", "已结算"))
            initial = {
                "store": row["store"],
                "report_type": row.get("report_type", "已结算"),
                "month_label": row["month_label"],
                "month_sort": row["month_sort"],
                "opening_balance": row["opening_balance"],
                "account_ending": row["account_ending"] if row["account_ending"] is not None else "0",
                "note": bal["note"] if bal else "",
            }
        elif self.current_store():
            initial = {"store": self.current_store(), "report_type": self.current_report_type_filter() or "已结算"}
        dialog = BalanceDialog(self, initial)
        if dialog.exec() == QDialog.Accepted:
            data = dialog.result_data
            self.repo.upsert_balance(data["store"], data["month_label"], data["month_sort"], data["opening_balance"], data["account_ending"], data["note"], data["report_type"])
            self.refresh_all()
            QMessageBox.information(self, "保存成功", "期初金额和店铺期末余额已保存。")

    def add_adjustment(self):
        row = self.selected_summary()
        if not row:
            QMessageBox.warning(self, "未选择", "请先选择一个店铺年月。")
            return
        dialog = AdjustmentDialog(self, row)
        if dialog.exec() == QDialog.Accepted:
            data = dialog.result_data
            self.repo.add_adjustment(row["store"], row["month_label"], row["month_sort"], data["target_column"], data["item"], data["amount"], data["note"], row.get("report_type", "已结算"))
            self.refresh_all()

    def show_difference(self):
        row = self.selected_summary()
        if not row:
            QMessageBox.warning(self, "未选择", "请先选择一个店铺年月。")
            return
        DifferenceDialog(self, self.repo, row).exec()

    def current_transaction_id(self):
        row = self.detail_table.currentRow()
        if row < 0 or not self.detail_table.item(row, 0):
            return None
        try:
            return int(self.detail_table.item(row, 0).text())
        except ValueError:
            return None

    def edit_transaction(self):
        tx_id = self.current_transaction_id()
        if not tx_id:
            QMessageBox.warning(self, "未选择", "请先在明细流水中选择一条记录。")
            return
        tx = self.repo.transaction_by_id(tx_id)
        if not tx:
            QMessageBox.warning(self, "未找到", "这条流水记录不存在，可能已被删除。")
            self.refresh_all()
            return
        initial = {
            "店铺": tx["store"],
            "月份显示": tx["month_label"],
            "月份排序(YYYY-MM)": tx["month_sort"],
            "动账时间": tx["transaction_time"],
            "动账方向": tx["direction"],
            "动账账户": tx["account"],
            "动账金额": money_text(tx["amount"]).replace(",", ""),
            "动账摘要": tx["summary"],
            "业务类型": tx["biz_type"],
        }
        dialog = TransactionDialog(self, initial)
        if dialog.exec() == QDialog.Accepted:
            self.repo.update_transaction(tx_id, dialog.result_data)
            self.refresh_all()

    def delete_transaction(self):
        tx_id = self.current_transaction_id()
        if not tx_id:
            QMessageBox.warning(self, "未选择", "请先在明细流水中选择一条记录。")
            return
        if QMessageBox.question(self, "确认删除", f"确定删除流水 ID {tx_id} 吗？") != QMessageBox.Yes:
            return
        self.repo.delete_transaction(tx_id)
        self.refresh_all()

    def delete_month(self):
        row = self.selected_summary()
        if not row:
            QMessageBox.warning(self, "未选择", "请先选择要删除的店铺年月。")
            return
        if QMessageBox.question(self, "确认删除", f"确定删除 {row['store']} {row['month_label']} 的全部流水和手工调整吗？余额记录默认保留。") != QMessageBox.Yes:
            return
        self.repo.delete_month(row["store"], row["month_sort"], delete_balance=False, report_type=row.get("report_type", "已结算"))
        self.refresh_all()

    def export_cell(self, worksheet, header, value):
        if value is None:
            return ""
        text = str(value)
        if text.lower() == "none":
            return ""
        if text.startswith("'"):
            text = text[1:]
        if ("订单编号" in header or header in ("主订单编号", "子订单编号")) and text.isdigit():
            cell = WriteOnlyCell(worksheet, value=int(text))
            cell.number_format = "0"
            return cell
        return text

    def export_summary(self):
        """导出汇总表和原始表格，原始表格按分页配置分批写入 XLSX。"""

        store = self.current_store()
        month_text = self.month_filter.text().strip() or "全部年月"
        type_text = self.current_report_type_filter() or "全部类型"
        export_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = "".join(ch if ch not in '\\/:*?"<>|' else "_" for ch in f"{store}_{type_text}_{month_text}_{export_time}")
        path, _ = QFileDialog.getSaveFileName(self, "保存汇总表", f"支付核对汇总_{safe_name}.xlsx", "Excel 文件 (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        if not store:
            QMessageBox.warning(self, "未选择店铺", "请先选择店铺。")
            return
        report_type = self.current_report_type_filter()
        rows = self.repo.monthly_summaries(store, self.month_filter.text().strip(), report_type_filter=report_type)
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(f"汇总_{store}_{month_text}"[:31])
        ws.title = f"汇总_{store}_{month_text}"[:31]
        headers = self.current_summary_columns()
        export_display_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append(["店铺", store])
        ws.append(["报表类型", report_type or "全部"])
        ws.append(["年月区间", month_text])
        ws.append(["导出时间", export_display_time])
        ws.append([])
        ws.append([display_header(header) for header in headers])
        for row in rows:
            ws.append([self.summary_value(row, header) for header in headers])
        raw_headers = self.repo.store_config(store)["raw_columns"]
        raw_ws = wb.create_sheet("原始表格")
        raw_ws.append(["店铺", "报表类型", "年月排序"] + [display_header(header) for header in raw_headers])
        total_raw = self.repo.raw_rows_count_for_export(store, self.month_filter.text().strip(), report_type_filter=report_type)
        batch_size = self.current_page_size()
        exported = 0
        while exported < total_raw:
            batch = self.repo.raw_rows_for_export_page(
                store,
                self.month_filter.text().strip(),
                report_type_filter=report_type,
                limit=batch_size,
                offset=exported,
            )
            for tx, payload in batch:
                raw_ws.append(
                    [tx["store"], tx["report_type"], tx["month_sort"]] +
                    [self.export_cell(raw_ws, header, payload.get(header, "")) for header in raw_headers]
                )
            exported += len(batch)
            self.set_status(f"正在导出原始表格：{exported}/{total_raw} 条")
            QApplication.processEvents()
            if not batch:
                break
        self.set_status("正在写入 XLSX 文件...")
        QApplication.processEvents()
        wb.save(path)
        self.set_status(f"导出完成：{Path(path).name}", path)
        message = QMessageBox(self)
        message.setIcon(QMessageBox.Information)
        message.setWindowTitle("导出完成")
        message.setText(f"已保存：{path}")
        open_btn = message.addButton("打开文件", QMessageBox.AcceptRole)
        message.addButton("确定", QMessageBox.RejectRole)
        message.exec()
        if message.clickedButton() == open_btn:
            os.startfile(path)


def main():
    app = QApplication(sys.argv)
    app.setApplicationName(f"{APP_TITLE} v{APP_VERSION}")
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
